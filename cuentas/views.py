from email.mime.application import MIMEApplication
from django.shortcuts import render, redirect
from django.contrib import messages
from .models import Usuario
from django.contrib.auth import logout
import openpyxl
from .models import Productos, Pedidos, Productos_4C, Pedidos_4C
from .models import Auditorias_HEB, Registro_Auditorias_HEB, Auditorias_HEB_Bebidas, Registro_Auditorias_HEB_Bebidas
from .models import Auditorias_4C, Registro_Auditorias_4C, Auditorias_4C_Bebidas,Registro_Auditorias_4C_Bebidas
from django.http import HttpResponse
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import datetime
import csv
# Variables globales para almacenar los nombres de los archivos generados
ultimo_pdf = ""
ultimo_txt = ""


def login_view(request):
    if request.method == 'POST':
        # Obtener los datos del formulario de inicio de sesión
        username = request.POST['username']
        password = request.POST['password']

        # Verificar el nombre de usuario y la contraseña en la base de datos
        try:
            usuario = Usuario.objects.get(username=username, password=password)
            # Nombre de usuario y contraseña son correctos
            messages.success(request, 'Inicio de sesión exitoso')
            request.session['login_usuario'] = usuario.id  # Guardar el id del usuario en la sesión
            return redirect('cuentas:menu', usuario_id=usuario.id)
        except Usuario.DoesNotExist:
            # Nombre de usuario o contraseña incorrectos
            messages.error(request, 'Usuario o contraseña incorrectos')

    return render(request, 'cuentas/login.html')


def registro_view(request):
    if request.method == 'POST':
        # Obtener los datos del formulario de registro
        nombres = request.POST['nombres']
        apellidos = request.POST['apellidos']
        username = request.POST['username']
        password = request.POST['password']
        sucursal = request.POST['sucursal']
        tipo_usuario = request.POST['tipo_usuario']

        # Guardar los datos en la base de datos
        usuario = Usuario(nombres=nombres, apellidos=apellidos, username=username, password=password, sucursal=sucursal, tipo_usuario=tipo_usuario)
        usuario.save()

        # Redirigir al usuario a una página de éxito o a otra página deseada
        return redirect('cuentas:login')

    return render(request, 'cuentas/registro.html')
    
def menu_view(request, usuario_id=None):
    if usuario_id is not None:
        usuario = Usuario.objects.get(id=usuario_id)
        # Obtener el nombre de la sucursal del usuario
        sucursal = usuario.sucursal
        tipo_usuario = usuario.tipo_usuario
        # Almacenar el nombre de la sucursal en la sesión
        request.session['sucursal'] = sucursal
    else:
        sucursal = ""  # No hay usuario autenticado
        tipo_usuario = ""  # No hay usuario autenticado

    context = {
        'sucursal': sucursal,
        'tipo_usuario': tipo_usuario,
    }

    if request.method == 'POST':
        if 'cerrarSesion' in request.POST:
            logout(request)  # Cerrar sesión
            return redirect('cuentas:login')
        elif 'actualizarBD' in request.POST:
            return redirect('cuentas:actualizar')
        elif 'pedidos' in request.POST:
            # Obtén el usuario_id de la sesión o establece un valor predeterminado si no existe
            usuario_id = request.session.get('login_usuario', None)
            if usuario_id is not None:
                return redirect('cuentas:pedidos', usuario_id=usuario_id)
            else:
                return redirect('cuentas:login')
        elif 'pedidos4C' in request.POST:
            usuario_id = request.session.get('login_usuario', None)
            if usuario_id is not None:
                return redirect('cuentas:pedidos4C', usuario_id=usuario_id)
            else:
                return redirect('cuentas:login')
        elif 'auditoria' in request.POST:
            return redirect('cuentas:auditoria')
        elif 'auditoriaBebidas' in request.POST:
            return redirect('cuentas:auditoriaBebidas')
        elif 'auditoria_4C' in request.POST:
            return redirect('cuentas:auditoria_4C')
        elif 'auditoria_4C_Bebidas' in request.POST:
            return redirect('cuentas:auditoria_4C_Bebidas')

    return render(request, 'cuentas/menu.html', context)



def actualizar_view(request):
    productos = Productos.objects.order_by('T1', 'DESCRIPCION')
    t1_choices = [
        (0.0, '0.0'),
        (1.0, '1.0'),
        (2.0, '2.0'),
        (3.0, '3.0'),
        (5.0, '5.0'),
        (5.5, '5.5'),
        (6.0, '6.0'),
        (7.0, '7.0'),
        (8.0, '8.0'),
        (8.5, '8.5'),
        (9.0, '9.0'),
        (9.5, '9.5'),
        (10.0, '10.0'),
        (10.5, '10.5'),
        (11.0, '11.0'),
        (12.0, '12.0'),
        (13.0, '13.0'),
        (16.0, '16.0'),
    ]

    kg_un_choices = [
        ("", "Selecciona"),
        ("kg", "kg"),
        ("pz", "pz")
    ]

    if request.method == 'POST':
        if 'archivo' in request.FILES:
            archivo = request.FILES['archivo']
            workbook = openpyxl.load_workbook(archivo)
            sheet = workbook.active

            productos_nuevos = []
            productos_eliminados = []

            for row in sheet.iter_rows(min_row=2, values_only=True):
                WRIN, DESCRIPCION = str(row[0]).strip().replace('="', '').replace('"', ''), row[1]
                if WRIN and DESCRIPCION:
                    producto, created = Productos.objects.get_or_create(
                        WRIN=WRIN,
                        defaults={
                            'T1': 0.0,
                            'TIPO': 'Caja',
                            'T2': 0.0,
                            'T3': 'Bolsa',
                            'CANT': 0,
                            'T4': 0.0,
                            'KG_UN': 'Selecciona',
                            'DESCRIPCION': DESCRIPCION,
                            'CJ': 0.0,
                            'BOLSA': 0.0,
                            'KG_PZ': 0.0,
                            'TOTAL': 0.0,
                            'ENVIAR': 0,  
                            'LIMIT': 0, 
                            'ACCESIBLE': 1,
                            'PEDIDO_ANTERIOR' : 0, 
                            'PROM_MENSUAL' : 0.0,                                             
                        }
                    )
                    if created:
                        productos_nuevos.append(producto)
                    else:
                        producto.DESCRIPCION = DESCRIPCION
                        producto.ACCESIBLE = 1
                        producto.save()

                    producto_4c, created_4c = Productos_4C.objects.get_or_create(
                        WRIN=WRIN,
                        defaults={
                            'T1': 0.0,
                            'TIPO': 'Caja',
                            'T2': 0.0,
                            'T3': 'Bolsa',
                            'CANT': 0,
                            'T4': 0.0,
                            'KG_UN': 'Selecciona',
                            'DESCRIPCION': DESCRIPCION,
                            'CJ': 0.0,
                            'BOLSA': 0.0,
                            'KG_PZ': 0.0,
                            'TOTAL': 0.0,
                            'ENVIAR': 0,  
                            'LIMIT': 0, 
                            'ACCESIBLE': 1,
                            'PEDIDO_ANTERIOR' : 0, 
                            'PROM_MENSUAL' : 0.0,  
                        }
                    )
                    if not created_4c:
                        producto_4c.DESCRIPCION = DESCRIPCION
                        producto_4c.ACCESIBLE = 1
                        producto_4c.save()

             # Eliminar productos seleccionados en Productos y CuentasProductos4C
            productos_eliminar = request.POST.getlist('eliminar')
            productos_inaccesibles = Productos.objects.filter(id__in=productos_eliminar)
            productos_inaccesibles.update(ACCESIBLE=0)

            productos_inaccesibles_4c = Productos_4C.objects.filter(id__in=productos_eliminar)
            productos_inaccesibles_4c.update(ACCESIBLE=0)

            # Eliminar productos que no están en el archivo Excel en Productos y CuentasProductos4C
            productos_en_excel = [str(row[0]).strip().replace('="', '').replace('"', '') for row in sheet.iter_rows(min_row=2, values_only=True)]
            productos_eliminar = Productos.objects.exclude(WRIN__in=productos_en_excel)
            productos_eliminar_list = list(productos_eliminar)
            productos_eliminar.update(ACCESIBLE=0)

            productos_eliminar_4c = Productos_4C.objects.exclude(WRIN__in=productos_en_excel)
            productos_eliminar_list_4c = list(productos_eliminar_4c)
            productos_eliminar_4c.update(ACCESIBLE=0)

            productos_eliminados = productos_eliminar_list

            return render(request, 'cuentas/exito.html', {'productos_nuevos': productos_nuevos, 'productos_eliminados': productos_eliminados})
        else:
            for producto in productos:
                t1 = request.POST.get(f'T1_{producto.id}')
                t2 = f'T2_{producto.id}'
                cant = f'CANT_{producto.id}'
                t4 = f'T4_{producto.id}'
                kg_un = request.POST.get(f'KG_UN_{producto.id}')
                limit = f'LIMIT_{producto.id}'
                if t1 is not None:
                    producto.T1 = float(t1)
                if cant in request.POST:
                    producto.CANT = int(request.POST[cant])
                if t4 in request.POST:
                    producto.T4 = float(request.POST[t4])
                cant = float(producto.CANT)
                t4 = float(producto.T4)
                producto.T2 = cant * t4 
                if t2 in request.POST:
                    producto.T2 = float(request.POST[t2])
                if kg_un is not None:
                    producto.KG_UN = kg_un
                if limit in request.POST:
                    producto.LIMIT = int(request.POST[limit])
                producto.save()

                producto_4c, created_4c = Productos_4C.objects.get_or_create(WRIN=producto.WRIN)
                producto_4c.T1 = producto.T1
                producto_4c.T2 = producto.T2
                producto_4c.CANT = producto.CANT
                producto_4c.T4 = producto.T4
                producto_4c.KG_UN = producto.KG_UN
                producto_4c.DESCRIPCION = producto.DESCRIPCION
                producto_4c.ACCESIBLE = producto.ACCESIBLE
                producto_4c.LIMIT = producto.LIMIT
                producto_4c.save()

            # Eliminar productos seleccionados
            productos_eliminar = request.POST.getlist('eliminar')
            productos_inaccesibles = Productos.objects.filter(id__in=productos_eliminar)
            productos_inaccesibles_4c = Productos_4C.objects.filter(id__in=productos_eliminar)
            productos_inaccesibles.update(ACCESIBLE=0)
            productos_inaccesibles_4c.update(ACCESIBLE=0)

            messages.success(request, 'Cambios guardados correctamente.')
            return redirect('cuentas:actualizar')

    return render(request, 'cuentas/actualizar.html', {'productos': productos, 't1_choices': t1_choices, 'kg_un_choices': kg_un_choices})


def exito_view(request):
    return render(request, 'cuentas/exito.html')

from django.db.models import Sum

def pedidos_view(request, usuario_id=None):
        
    productos = Productos.objects.all().order_by('T1','DESCRIPCION')

    context = {}
    if request.method == 'POST':
        if 'ir_a_menu' in request.POST:
            usuario_id = request.session.get('login_usuario', None)
            if usuario_id is not None:
                context['usuario_id'] = usuario_id
                return redirect('cuentas:menu', usuario_id = usuario_id)
        # Recuperar los productos de la base de datos
        # Actualizar los valores de los campos Cant_a_enviar e inventario
        for producto in productos:
            cj_key = f'cj_{producto.id}'
            bolsa_key = f'bolsa_{producto.id}'
            kg_pz_key = f'kg_pz_{producto.id}'
            total = f'total_{producto.id}'
            enviar_key = f'enviar_{producto.id}'
            if cj_key in request.POST:
                producto.CJ = float(request.POST[cj_key])
            if bolsa_key in request.POST:
                producto.BOLSA = float(request.POST[bolsa_key])
            if kg_pz_key in request.POST:
                producto.KG_PZ = float(request.POST[kg_pz_key])
            
            # Calcular el valor de TOTAL
            cj = float(producto.CJ)
            bolsa = float(producto.BOLSA)
            kg_pz = float(producto.KG_PZ)
            producto.TOTAL = cj * producto.T2 + bolsa * producto.T4 + kg_pz
            if total in request.POST:
                producto.TOTAL = float(request.POST[total])
            if enviar_key in request.POST:
                producto.ENVIAR = float(request.POST[enviar_key])
            producto.save()

        # Calcular la suma total de CANT_A_ENVIAR
        suma_total = Productos.objects.aggregate(total=Sum('ENVIAR')).get('total')

        # Pasar los productos y la suma total a la plantilla
        context = {
            'productos': productos,
            'suma_total': suma_total,
        }

        # Redireccionar a una página de éxito o a otra vista si es necesario
        return redirect('cuentas:pedidos')
    else:
        # Recuperar los productos de la base de datos

        # Pasar los productos a la plantilla
        context = {
            'productos': productos,
        }
        return render(request, 'cuentas/pedidos.html', context)

def pedidos4C_view(request, usuario_id=None):
    productos = Productos_4C.objects.all().order_by('T1','DESCRIPCION')

    context = {}
    if request.method == 'POST':
        if 'ir_a_menu' in request.POST:
            usuario_id = request.session.get('login_usuario', None)
            if usuario_id is not None:
                context['usuario_id'] = usuario_id
                return redirect('cuentas:menu', usuario_id = usuario_id)
        # Recuperar los productos de la base de datos

        # Actualizar los valores de los campos Cant_a_enviar e inventario
        for producto in productos:
            cj_key = f'cj_{producto.id}'
            bolsa_key = f'bolsa_{producto.id}'
            kg_pz_key = f'kg_pz_{producto.id}'
            total = f'total_{producto.id}'
            enviar_key = f'enviar_{producto.id}'
            if cj_key in request.POST:
                producto.CJ = float(request.POST[cj_key])
            if bolsa_key in request.POST:
                producto.BOLSA = float(request.POST[bolsa_key])
            if kg_pz_key in request.POST:
                producto.KG_PZ = float(request.POST[kg_pz_key])
            
            # Calcular el valor de TOTAL
            cj = float(producto.CJ)
            bolsa = float(producto.BOLSA)
            kg_pz = float(producto.KG_PZ)
            producto.TOTAL = cj * producto.T2 + bolsa * producto.T4 + kg_pz
            if total in request.POST:
                producto.TOTAL = float(request.POST[total])
            if enviar_key in request.POST:
                producto.ENVIAR = float(request.POST[enviar_key])
            producto.save()

        # Calcular la suma total de CANT_A_ENVIAR
        suma_total = Productos_4C.objects.aggregate(total=Sum('ENVIAR')).get('total')

        # Pasar los productos y la suma total a la plantilla
        context = {
            'productos': productos,
            'suma_total': suma_total,
        }

        # Redireccionar a una página de éxito o a otra vista si es necesario
        return redirect('cuentas:pedidos4C')
    else:
        # Recuperar los productos de la base de datos

        # Pasar los productos a la plantilla
        context = {
            'productos': productos,
        }
        return render(request, 'cuentas/pedidos_4C.html', context)

import datetime

def exportar_pdf(request):
    response = HttpResponse(content_type='application/pdf')

    # Obtener el id del usuario de la sesión
    login_usuario_id = request.session.get('login_usuario')

    if login_usuario_id is not None:
        try:
            usuario = Usuario.objects.get(id=login_usuario_id)
            # Obtener los datos del usuario
            nombres = usuario.nombres
            apellidos = usuario.apellidos
            sucursal = usuario.sucursal

            # Construir el nombre del archivo según la sucursal
            if sucursal == 'HEB':
                productos = Productos.objects.filter(ENVIAR__gt=0).order_by('T1','DESCRIPCION')
                nombre_archivo = f'Pedido subway 60909 {datetime.datetime.now().strftime("%d-%m-%Y-%H-%M-%S")}.pdf'
            elif sucursal == '4 Caminos':
                productos = Productos_4C.objects.filter(ENVIAR__gt=0).order_by('T1','DESCRIPCION')
                nombre_archivo = f'Pedido subway 33103 {datetime.datetime.now().strftime("%d-%m-%Y-%H-%M-%S")}.pdf'
            else:
                nombre_archivo = 'exportacion.pdf'

            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

            data = [['DESCRIPCION', 'CJ', 'BOLSA', 'KG_PZ','TOTAL', 'KG_UN','ENVIAR']]

            for producto in productos:
                data.append([
                    str(producto.DESCRIPCION),
                    str(producto.CJ),
                    str(producto.BOLSA),
                    str(producto.KG_PZ),
                    str(producto.TOTAL),
                    str(producto.KG_UN),
                    str(producto.ENVIAR),
                ])

            doc = SimpleDocTemplate(response, pagesize=letter)
            elements = []

            # Agregar encabezado al PDF
            styles = getSampleStyleSheet()
            elements.append(Paragraph(f'Bienvenidos a SUBWAY {sucursal}', styles['Heading1']))
            elements.append(Paragraph(f'Pedido elaborado por: {nombres} {apellidos}', styles['Normal']))
            
            # Agregar fecha y hora de generación
            fecha_generacion = datetime.datetime.now().strftime('%d-%m-%Y %H:%M:%S')
            elements.append(Spacer(1, 12))
            elements.append(Paragraph(f'Fecha y hora de generación: {fecha_generacion}', styles['Normal']))
            elements.append(Spacer(1, 12))
            
            # Calcular la suma del campo 'CANT_A_ENVIAR'
            suma_total = sum([int(producto.ENVIAR) for producto in productos])
            
            # Agregar la suma debajo de la fecha y hora
            elements.append(Spacer(1, 12))
            elements.append(Paragraph('Cantidad a enviar total: ' + str(suma_total), styles['Normal']))
            
            filtered_data = [[cell for cell in row] for row in data if any(cell != '0' for cell in row[4:6])]
            filtered_table = Table(filtered_data)

            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.green),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ])
            
            filtered_table.setStyle(style)
            elements.append(filtered_table)

            doc.build(elements)

            return response
        except Usuario.DoesNotExist:
            pass  # Manejar la situación si el usuario no existe

    # En caso de no haber usuario autenticado o error en la consulta
    # redireccionar a la página de inicio de sesión u otra página apropiada
    return response


def exportar_txt(request):
    response = HttpResponse(content_type='text/plain')

    login_usuario_id = request.session.get('login_usuario')
    if login_usuario_id is not None:
        try:
            usuario = Usuario.objects.get(id=login_usuario_id)
            # Obtener los datos del usuario
            nombres = usuario.nombres
            apellidos = usuario.apellidos
            sucursal = usuario.sucursal

            # Construir el nombre del archivo según la sucursal
            if sucursal == 'HEB':
                productos = Productos.objects.filter(ACCESIBLE=1)
                nombre_archivo = f'Pedido subway 60909 {datetime.datetime.now().strftime("%d-%m-%Y-%H-%M-%S")}.txt'
            elif sucursal == '4 Caminos':
                productos = Productos_4C.objects.filter(ACCESIBLE=1)
                nombre_archivo = f'Pedido subway 33103 {datetime.datetime.now().strftime("%d-%m-%Y-%H-%M-%S")}.txt'
            else:
                nombre_archivo = 'exportacion.txt'

            response['Content-Disposition'] = f'attachment; filename="{nombre_archivo}"'

            writer = csv.writer(response, delimiter='\t')

            # Encabezado del archivo de texto
            writer.writerow(['Código/WRIN', 'Descripción', 'Rubro', 'Codigo Micelaneo', 'Presentación', ' ', 'Cant. a enviar'])

            # Contenido del archivo de texto
            for producto in productos:
                writer.writerow([producto.WRIN, producto.DESCRIPCION, ' ', ' ', ' ', ' ', producto.ENVIAR])

            return response
        except Usuario.DoesNotExist:
            pass  # Manejar la situación si el usuario no existe

    # Si no se encuentra un usuario en sesión o hay un error, se devuelve una respuesta vacía
    return response

def enviar_correo(request):
    # Obtener el id del usuario de la sesión
    login_usuario_id = request.session.get('login_usuario')

    if login_usuario_id is not None:
        try:
            usuario = Usuario.objects.get(id=login_usuario_id)
            # Obtener el valor de sucursal del campo correspondiente en la tabla cuentas_usuario
            sucursal = usuario.sucursal
        except Usuario.DoesNotExist:
            sucursal = 'NombreDeSucursalPorDefecto'  # Manejar la situación si el usuario no existe
    else:
        sucursal = 'NombreDeSucursalPorDefecto'  # Manejar la situación si no hay un usuario en sesión

    # Construir el encabezado del correo electrónico
    subject = f'Pedido Subway {sucursal} {datetime.datetime.now().strftime("%Y-%m-%d %H-%M-%S")}'

    # Construir el destinatario del correo electrónico
    to_email = 'vicente@grupoelma.com'

    # Construir el enlace de correo electrónico con los parámetros predefinidos
    mailto_link = f'mailto:{to_email}?subject={subject}'

    # Obtener la URL de la página anterior
    referer = request.META.get('HTTP_REFERER')

    # Crear el botón para enviar el correo electrónico
    button_html = f'<a href="{mailto_link}">Enviar correo electrónico</a>'

    # Crear el botón para regresar a la página anterior
    back_button_html = f'<a href="{referer}">Regresar</a>'

    # Combinar ambos botones en una sola respuesta HTTP
    response_html = f'{button_html} | {back_button_html}'

    return HttpResponse(response_html)


from django.shortcuts import get_object_or_404

def eliminar_producto(request, producto_id):
    producto = get_object_or_404(Productos, id=producto_id)

    # Eliminar el producto de ambos modelos, si existe en alguno de ellos
    if request.method == 'POST':
        confirmacion = request.POST.get('confirmar', 'no')
        if confirmacion == 'sí':
            # Verificar si el producto existe en el modelo CuentasProductos4C
            try:
                producto_4c = Productos_4C.objects.get(WRIN=producto.WRIN)
                producto_4c.delete()
            except Productos_4C.DoesNotExist:
                pass

            # Eliminar el producto del modelo Productos
            producto.delete()

    return redirect('cuentas:actualizar')

from django.db import models
from django.db.models import Max, F, Sum

def finalizar_pedido(request):
   # Obtener el último valor de 'NO_PEDIDO' y sumarle 1
    numero_pedido_actual = Pedidos.objects.aggregate(max_pedido=Max('NO_PEDIDO'))['max_pedido']
    if numero_pedido_actual is None:
        numero_pedido_actual = 1
    else:
        numero_pedido_actual += 1


    # Obtener los productos con ENVIAR diferente de 0
    productos_a_guardar = Productos.objects.filter(ENVIAR__gt=0)

     # Actualizar el campo 'PEDIDO_ANTERIOR' de los productos en la tabla 'cuentas_productos'
    # Establecer a 0 todos los valores que no están en 'productos_a_guardar'
    Productos.objects.exclude(id__in=productos_a_guardar).update(PEDIDO_ANTERIOR=0)

    # Actualizar el campo 'PEDIDO_ANTERIOR' de los productos en la tabla 'cuentas_productos'
    productos_a_guardar.update(PEDIDO_ANTERIOR=models.F('ENVIAR'))

    # Guardar los productos en la tabla 'cuentas_pedidos'
    for producto in productos_a_guardar:
        Pedidos.objects.create(
            WRIN=producto.WRIN,
            DESCRIPCION=producto.DESCRIPCION,
            ENVIAR=producto.ENVIAR,
            NO_PEDIDO=numero_pedido_actual  # Asignar el número de pedido actual al campo NO_PEDIDO
        )
    
    # Obtener los últimos 4 pedidos que tienen el mismo WRIN en la tabla 'cuentas_pedidos'
        if numero_pedido_actual >= 4:
            pedidos_mismo_wrin = Pedidos.objects.filter(WRIN=producto.WRIN).order_by('-NO_PEDIDO')[:4]
        else:
            pedidos_mismo_wrin = Pedidos.objects.filter(WRIN=producto.WRIN)


        # Calcular el PROM_MENSUAL para los productos correspondientes
        enviar_total = pedidos_mismo_wrin.aggregate(total_enviar=Sum('ENVIAR'))['total_enviar']
        promedio_mensual = round(enviar_total / pedidos_mismo_wrin.count(), 2)

        # Actualizar el campo PROM_MENSUAL en 'cuentas_productos' para el WRIN correspondiente
        Productos.objects.filter(WRIN=producto.WRIN).update(PROM_MENSUAL=promedio_mensual)

    # Redirigir a la página de pedidos
    return redirect('cuentas:pedidos')

def finalizar_pedido4C(request):
   # Obtener el último valor de 'NO_PEDIDO' y sumarle 1
    numero_pedido_actual = Pedidos_4C.objects.aggregate(max_pedido=Max('NO_PEDIDO'))['max_pedido']
    if numero_pedido_actual is None:
        numero_pedido_actual = 1
    else:
        numero_pedido_actual += 1


    # Obtener los productos con ENVIAR diferente de 0
    productos_a_guardar = Productos_4C.objects.filter(ENVIAR__gt=0)

     # Actualizar el campo 'PEDIDO_ANTERIOR' de los productos en la tabla 'cuentas_productos'
    # Establecer a 0 todos los valores que no están en 'productos_a_guardar'
    Productos_4C.objects.exclude(id__in=productos_a_guardar).update(PEDIDO_ANTERIOR=0)

    # Actualizar el campo 'PEDIDO_ANTERIOR' de los productos en la tabla 'cuentas_productos'
    productos_a_guardar.update(PEDIDO_ANTERIOR=models.F('ENVIAR'))

    # Guardar los productos en la tabla 'cuentas_pedidos'
    for producto in productos_a_guardar:
        Pedidos_4C.objects.create(
            WRIN=producto.WRIN,
            DESCRIPCION=producto.DESCRIPCION,
            ENVIAR=producto.ENVIAR,
            NO_PEDIDO=numero_pedido_actual  # Asignar el número de pedido actual al campo NO_PEDIDO
        )
    
    # Obtener los últimos 4 pedidos que tienen el mismo WRIN en la tabla 'cuentas_pedidos'
        if numero_pedido_actual >= 4:
            pedidos_mismo_wrin = Pedidos_4C.objects.filter(WRIN=producto.WRIN).order_by('-NO_PEDIDO')[:4]
        else:
            pedidos_mismo_wrin = Pedidos_4C.objects.filter(WRIN=producto.WRIN)


        # Calcular el PROM_MENSUAL para los productos correspondientes
        enviar_total = pedidos_mismo_wrin.aggregate(total_enviar=Sum('ENVIAR'))['total_enviar']
        promedio_mensual = round(enviar_total / pedidos_mismo_wrin.count(), 2)

        # Actualizar el campo PROM_MENSUAL en 'cuentas_productos' para el WRIN correspondiente
        Productos_4C.objects.filter(WRIN=producto.WRIN).update(PROM_MENSUAL=promedio_mensual)

    # Redirigir a la página de pedidos
    return redirect('cuentas:pedidos4C')

def auditoria(request):
    productos = Auditorias_HEB.objects.all()

    # Generar una lista de días del 1 al 31
    dias = list(range(1, 32))

    # Generar una lista de años desde 2020 hasta 2050
    años = list(range(2023, 2051))
    if request.method == 'POST' and 'limpiar_HEB' in request.POST:

        for producto in productos:
            producto.DIA = 0
            producto.MES = 'Mes'
            producto.AÑO = 0 
            producto.DIA_SEMANAL = 'Dia'
            producto.DIA_ANTERIOR = 0.0
            producto.DIA_ACTUAL = 0.0
            producto.CAJAS = 0
            producto.COMPRAS = 0.0
            producto.ENTRADA = 0.0
            producto.SALIDA = 0.0
            producto.MERMAS = 0.0
            producto.CONSUMO_DIARIO = 0.0
            producto.save()
        return render(request, 'cuentas/auditoria.html', {'productos': productos, 'dias': dias, 'años': años})

    if request.method == 'POST':
        if 'archivo' in request.FILES:
            archivo = request.FILES['archivo']
            workbook = openpyxl.load_workbook(archivo)
            sheet = workbook.active

            productos_nuevos = []

            for row in sheet.iter_rows(min_row=4, values_only=True):
                DESCRIPCION, MEDIDA = row[1], str(row[2])

                if DESCRIPCION and MEDIDA:
                    producto, created = Auditorias_HEB.objects.get_or_create(
                        DESCRIPCION=DESCRIPCION,
                        defaults={
                            'MEDIDA': MEDIDA,
                            'DIA_ANTERIOR': 0.0,
                            'DIA_ACTUAL': 0.0,
                            'CAJAS': 0,
                            'COMPRAS': 0.0,
                            'ENTRADA': 0.0,
                            'SALIDA': 0.0,
                            'MERMAS': 0.0,
                            'CONSUMO_DIARIO': 0.0,
                            'DIA_SEMANAL': 'Dia',
                            'DIA': 0,
                            'MES': 'Mes',
                            'AÑO': 0,
                        }
                    )
                    if created:
                        productos_nuevos.append(producto)
                    else:
                        producto.DESCRIPCION = DESCRIPCION
                        producto.MEDIDA = MEDIDA
                        producto.save()
    
    if request.method == 'POST' and 'guardar_cambios' in request.POST:
        # Procesar la selección de DÍA, MES y AÑO
        dia_seleccionado = int(request.POST.get('dia', 0))
        mes_seleccionado = request.POST.get('mes', '')
        año_seleccionado = int(request.POST.get('anio', 0))
        dia_semanal_seleccionado = request.POST.get('dia_semanal', '')
        

        # Iterar sobre los productos y actualizar los campos correspondientes
        for producto in productos:
            dia_anterior = float(request.POST.get(f'dia_anterior_{producto.id}', 0))
            dia_actual = float(request.POST.get(f'dia_actual_{producto.id}', 0))
            cajas = int(request.POST.get(f'cajas_{producto.id}', 0))
            compras = float(request.POST.get(f'compras_{producto.id}', 0))
            entrada = float(request.POST.get(f'entrada_{producto.id}', 0))
            salida = float(request.POST.get(f'salida_{producto.id}', 0))
            mermas = float(request.POST.get(f'mermas_{producto.id}', 0))

            # Actualizar los campos del producto
            producto.DIA = dia_seleccionado
            producto.MES = mes_seleccionado
            producto.AÑO = año_seleccionado
            producto.DIA_SEMANAL = dia_semanal_seleccionado
            producto.DIA_ANTERIOR = dia_anterior
            producto.DIA_ACTUAL = dia_actual
            producto.CAJAS = cajas
            if producto.id == 1:
                producto.COMPRAS = producto.CAJAS * 70
            elif producto.id == 2:
                producto.COMPRAS = producto.CAJAS * 9.6
            elif producto.id == 3:
                producto.COMPRAS = producto.CAJAS * 11.5
            elif producto.id == 4:
                producto.COMPRAS = producto.CAJAS * 8
            elif producto.id == 5:
                producto.COMPRAS = producto.CAJAS * 7
            elif producto.id == 6:
                producto.COMPRAS = producto.CAJAS * 10
            elif producto.id == 7:
                producto.COMPRAS = producto.CAJAS * 9.6
            elif producto.id == 8:
                producto.COMPRAS = producto.CAJAS * 6
            elif producto.id == 9:
                producto.COMPRAS = 0.0
            elif producto.id == 10:
                producto.COMPRAS = producto.CAJAS * 12
            elif producto.id == 11:
                producto.COMPRAS = producto.CAJAS * 75
            elif producto.id == 12:
                producto.COMPRAS = producto.CAJAS * 600
            elif producto.id == 13:
                producto.COMPRAS = producto.CAJAS * 7
            elif producto.id == 14:
                producto.COMPRAS = producto.CAJAS * 9.08
            elif producto.id == 15:
                producto.COMPRAS = producto.CAJAS * 9.08
            elif producto.id == 16:
                producto.COMPRAS = producto.CAJAS * 9
            elif producto.id == 17:
                producto.COMPRAS = producto.CAJAS * 8
            elif producto.id == 18:
                producto.COMPRAS = producto.CAJAS * 8
            elif producto.id == 19:
                producto.COMPRAS = producto.CAJAS * 8
            elif producto.id == 20:
                producto.COMPRAS = producto.CAJAS * 8
            elif producto.id == 21:
                producto.COMPRAS = producto.CAJAS * 8
            elif producto.id == 22:
                producto.COMPRAS = producto.CAJAS * 8
            elif producto.id == 23:
                producto.COMPRAS = producto.CAJAS * 4
            elif producto.id == 24:
                producto.COMPRAS = producto.CAJAS * 8
            elif producto.id == 25:
                producto.COMPRAS = producto.CAJAS * 4
            elif producto.id == 26:
                producto.COMPRAS = producto.CAJAS * 6
            elif producto.id == 27:
                producto.COMPRAS = producto.CAJAS * 6
            elif producto.id == 28:
                producto.COMPRAS = producto.CAJAS * 50
            elif producto.id == 29:
                producto.COMPRAS = producto.CAJAS * 200
            elif producto.id == 30:
                producto.COMPRAS = producto.CAJAS
            elif producto.id == 31:
                producto.COMPRAS = 0.0
            producto.ENTRADA = entrada
            producto.SALIDA = salida
            producto.MERMAS = mermas

            # Calcular el consumo diario
            if dia_actual == 0:
                producto.CONSUMO_DIARIO = 0
            else:
                producto.CONSUMO_DIARIO = (-dia_anterior) - producto.COMPRAS + dia_actual - entrada + salida - mermas

            producto.save()

            if dia_semanal_seleccionado == 'Viernes':
                    # Verificar si existe un registro en Registro_Auditorias_HEB con la misma fecha
                registro_existente = Registro_Auditorias_HEB.objects.filter(
                    DESCRIPCION=producto.DESCRIPCION,
                    MEDIDA=producto.MEDIDA,
                    DIA_SEMANAL = dia_semanal_seleccionado,
                    DIA=dia_seleccionado,
                    MES=mes_seleccionado,
                    AÑO=año_seleccionado
                    ).first()
                
                if dia_seleccionado == 1:
                    if mes_seleccionado == 'Septiembre' and año_seleccionado == 2023:
                        registro_anterior = Registro_Auditorias_HEB.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=31, 
                        MES='Agosto',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Octubre' and año_seleccionado == 2023:
                        registro_anterior = Registro_Auditorias_HEB.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=30, 
                        MES='Septiembre',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Noviembre' and año_seleccionado == 2023:
                        registro_anterior = Registro_Auditorias_HEB.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=31, 
                        MES='Octubre',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Diciembre' and año_seleccionado == 2023:
                        registro_anterior = Registro_Auditorias_HEB.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=30, 
                        MES='Noviembre',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Enero' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_HEB.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=31, 
                        MES='Diciembre',
                        AÑO=2023
                        ).first()
                    elif mes_seleccionado == 'Febrero' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_HEB.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=31, 
                        MES='Enero',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Marzo' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_HEB.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=28, 
                        MES='Febrero',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Abril' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_HEB.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=31, 
                        MES='Marzo',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Mayo' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_HEB.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=30, 
                        MES='Abril',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Junio' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_HEB.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=31, 
                        MES='Mayo',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Julio' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_HEB.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=30, 
                        MES='Junio',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Agosto' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_HEB.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=31, 
                        MES='Julio',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Septiembre' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_HEB.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=31, 
                        MES='Agosto',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Octubre' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_HEB.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=30, 
                        MES='Septiembre',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Noviembre' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_HEB.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=31, 
                        MES='Octubre',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Diciembre' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_HEB.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=30, 
                        MES='Noviembre',
                        AÑO=año_seleccionado
                        ).first()
                else:
                    registro_anterior = Registro_Auditorias_HEB.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=dia_seleccionado - 1, 
                        MES=mes_seleccionado,
                        AÑO=año_seleccionado
                        ).first()

                if registro_existente:
                    checkbox_name = f'check_{producto.id}'  # Nombre del campo checkbox
                    if request.POST.get(checkbox_name) == 'on':
                        producto.DIA_ANTERIOR = round(producto.DIA_ANTERIOR - producto.COMPRAS, 3)
                        producto.CONSUMO_DIARIO = (-producto.DIA_ANTERIOR) - producto.COMPRAS + dia_actual - entrada + salida - mermas
                        producto.save()
                        registro_anterior.DIA_ACTUAL = producto.DIA_ANTERIOR
                        registro_anterior.CONSUMO_DIARIO = (-registro_anterior.DIA_ANTERIOR) - registro_anterior.COMPRAS + registro_anterior.DIA_ACTUAL - registro_anterior.ENTRADA + registro_anterior.SALIDA - registro_anterior.MERMAS
                        registro_anterior.save()
                    else:
                        if registro_anterior:
                            registro_anterior.DIA_ACTUAL = producto.DIA_ANTERIOR
                            registro_anterior.CONSUMO_DIARIO = (-registro_anterior.DIA_ANTERIOR) - registro_anterior.COMPRAS + registro_anterior.DIA_ACTUAL - registro_anterior.ENTRADA + registro_anterior.SALIDA - registro_anterior.MERMAS
                            registro_anterior.save()
                 # Si existe un registro con la misma fecha, actualizar los datos en Registro_Auditorias_HEB
                    registro_existente.DIA_ANTERIOR = producto.DIA_ANTERIOR 
                    registro_existente.DIA_ACTUAL = producto.DIA_ACTUAL
                    registro_existente.CAJAS = producto.CAJAS
                    registro_existente.COMPRAS = producto.COMPRAS
                    registro_existente.ENTRADA = producto.ENTRADA
                    registro_existente.SALIDA = producto.SALIDA
                    registro_existente.MERMAS = producto.MERMAS
                    registro_existente.CONSUMO_DIARIO = producto.CONSUMO_DIARIO
                    registro_existente.save()
                else:
            # Si no existe un registro con la misma fecha, insertar un nuevo registro en Registro_Auditorias_HEB
                    checkbox_name = f'check_{producto.id}'  # Nombre del campo checkbox
                    if request.POST.get(checkbox_name) == 'on':
                        producto.DIA_ANTERIOR = round(producto.DIA_ANTERIOR - producto.COMPRAS, 3)
                        producto.CONSUMO_DIARIO = (-producto.DIA_ANTERIOR) - producto.COMPRAS + dia_actual - entrada + salida - mermas
                        producto.save()
                        registro_anterior.DIA_ACTUAL = producto.DIA_ANTERIOR
                        registro_anterior.CONSUMO_DIARIO = (-registro_anterior.DIA_ANTERIOR) - registro_anterior.COMPRAS + registro_anterior.DIA_ACTUAL - registro_anterior.ENTRADA + registro_anterior.SALIDA - registro_anterior.MERMAS
                        registro_anterior.save()
                    else:
                        if registro_anterior:
                            registro_anterior.DIA_ACTUAL = producto.DIA_ANTERIOR
                            registro_anterior.CONSUMO_DIARIO = (-registro_anterior.DIA_ANTERIOR) - registro_anterior.COMPRAS + registro_anterior.DIA_ACTUAL - registro_anterior.ENTRADA + registro_anterior.SALIDA - registro_anterior.MERMAS
                            registro_anterior.save()
                    nuevo_registro = Registro_Auditorias_HEB(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA_ANTERIOR=producto.DIA_ANTERIOR,
                        DIA_ACTUAL=producto.DIA_ACTUAL,
                        CAJAS=producto.CAJAS,
                        COMPRAS=producto.COMPRAS,
                        ENTRADA=producto.ENTRADA,
                        SALIDA=producto.SALIDA,
                        MERMAS=producto.MERMAS,
                        CONSUMO_DIARIO=producto.CONSUMO_DIARIO,
                        DIA_SEMANAL = dia_semanal_seleccionado,
                        DIA=dia_seleccionado,
                        MES=mes_seleccionado,
                        AÑO=año_seleccionado
                    )
                    nuevo_registro.save()
                    
            else:
                    # Verificar si existe un registro en Registro_Auditorias_HEB con la misma fecha
                registro_existente = Registro_Auditorias_HEB.objects.filter(
                    DESCRIPCION=producto.DESCRIPCION,
                    MEDIDA=producto.MEDIDA,
                    DIA_SEMANAL = dia_semanal_seleccionado,
                    DIA=dia_seleccionado,
                    MES=mes_seleccionado,
                    AÑO=año_seleccionado
                    ).first()

                if registro_existente:
                 # Si existe un registro con la misma fecha, actualizar los datos en Registro_Auditorias_HEB
                    registro_existente.DIA_ANTERIOR = producto.DIA_ANTERIOR
                    registro_existente.DIA_ACTUAL = producto.DIA_ACTUAL
                    registro_existente.CAJAS = producto.CAJAS
                    registro_existente.COMPRAS = producto.COMPRAS
                    registro_existente.ENTRADA = producto.ENTRADA
                    registro_existente.SALIDA = producto.SALIDA
                    registro_existente.MERMAS = producto.MERMAS
                    registro_existente.CONSUMO_DIARIO = producto.CONSUMO_DIARIO
                    registro_existente.save()
                else:
            # Si no existe un registro con la misma fecha, insertar un nuevo registro en Registro_Auditorias_HEB
                    nuevo_registro = Registro_Auditorias_HEB(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA_ANTERIOR=producto.DIA_ANTERIOR,
                        DIA_ACTUAL=producto.DIA_ACTUAL,
                        CAJAS=producto.CAJAS,
                        COMPRAS=producto.COMPRAS,
                        ENTRADA=producto.ENTRADA,
                        SALIDA=producto.SALIDA,
                        MERMAS=producto.MERMAS,
                        CONSUMO_DIARIO=producto.CONSUMO_DIARIO,
                        DIA_SEMANAL = dia_semanal_seleccionado,
                        DIA=dia_seleccionado,
                        MES=mes_seleccionado,
                        AÑO=año_seleccionado
                        )
                    nuevo_registro.save()
            
                
        return render(request, 'cuentas/auditoria.html', {
            'dias': dias,
            'años': años,
            'productos': productos,
            'dia_semanal_seleccionado': dia_semanal_seleccionado,
            'dia_seleccionado': dia_seleccionado,
            'mes_seleccionado': mes_seleccionado,
            'año_seleccionado': año_seleccionado,
        })

    return render(request, 'cuentas/auditoria.html', {'productos': productos, 'dias': dias, 'años': años})


def recuperar_datos(request):
     # Generar una lista de días del 1 al 31
    dias = list(range(1, 32))

    # Generar una lista de años desde 2020 hasta 2050
    años = list(range(2023, 2051))
    if request.method == 'POST':
        # Obtener el día, mes y año seleccionados
        dia_seleccionado = int(request.POST.get('dia', 0))
        mes_seleccionado = request.POST.get('mes', '')
        año_seleccionado = int(request.POST.get('anio', 0))
        dia_semanal_seleccionado = request.POST.get('dia_semanal', '')
        productos = Auditorias_HEB.objects.all()
        
        for producto in productos:
            registro = Registro_Auditorias_HEB.objects.filter(
                DESCRIPCION=producto.DESCRIPCION,
                MEDIDA=producto.MEDIDA,
                DIA=dia_seleccionado,
                MES=mes_seleccionado,
                AÑO=año_seleccionado
            ).first()
            
            if registro:
                # Si existen registros, actualizar los datos en Auditorias_HEB
                producto.DIA_ANTERIOR = registro.DIA_ANTERIOR
                producto.DIA_ACTUAL = registro.DIA_ACTUAL
                producto.CAJAS = registro.CAJAS
                producto.COMPRAS = registro.COMPRAS
                producto.ENTRADA = registro.ENTRADA
                producto.SALIDA = registro.SALIDA
                producto.MERMAS = registro.MERMAS
                producto.CONSUMO_DIARIO = registro.CONSUMO_DIARIO
                producto.DIA_SEMANAL = registro.DIA_SEMANAL
                producto.DIA = registro.DIA
                producto.MES = registro.MES
                producto.AÑO = registro.AÑO
                producto.save()
            else:
                if dia_seleccionado == 1 and año_seleccionado == 2023:
                    if mes_seleccionado == 'Septiembre':
                        registro_anterior = Registro_Auditorias_HEB.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=31, 
                            MES='Agosto',
                            AÑO=año_seleccionado
                            ).first()
                    elif mes_seleccionado == 'Octubre':
                        registro_anterior = Registro_Auditorias_HEB.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=30, 
                            MES='Septiembre',
                            AÑO=año_seleccionado
                            ).first()
                    elif mes_seleccionado == 'Noviembre':
                        registro_anterior = Registro_Auditorias_HEB.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=31, 
                            MES='Octubre',
                            AÑO=año_seleccionado
                            ).first()
                    elif mes_seleccionado == 'Diciembre':
                        registro_anterior = Registro_Auditorias_HEB.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=30, 
                            MES='Noviembre',
                            AÑO=año_seleccionado
                            ).first()
                    if registro_anterior:
                        producto.DIA_ANTERIOR = registro_anterior.DIA_ACTUAL
                        producto.DIA_ACTUAL = 0.0
                        producto.CAJAS = 0
                        producto.COMPRAS = 0.0
                        producto.ENTRADA = 0.0
                        producto.SALIDA = 0.0
                        producto.MERMAS = 0.0
                        producto.save()
                else:
                    if dia_seleccionado == 1 and año_seleccionado == 2024:
                        if mes_seleccionado == 'Enero':
                            registro_anterior = Registro_Auditorias_HEB.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Diciembre',
                                AÑO=2023
                                ).first()
                        if mes_seleccionado == 'Febrero':
                            registro_anterior = Registro_Auditorias_HEB.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Enero',
                                AÑO=año_seleccionado
                                ).first()
                        elif mes_seleccionado == 'Marzo':
                            registro_anterior = Registro_Auditorias_HEB.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=28, 
                                MES='Febrero',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Abril':
                            registro_anterior = Registro_Auditorias_HEB.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=31, 
                            MES='Marzo',
                            AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Mayo':
                            registro_anterior = Registro_Auditorias_HEB.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=30, 
                                MES='Abril',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Junio':
                            registro_anterior = Registro_Auditorias_HEB.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Mayo',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Julio':
                            registro_anterior = Registro_Auditorias_HEB.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=30, 
                                MES='Junio',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Agosto':
                            registro_anterior = Registro_Auditorias_HEB.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Julio',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Septiembre':
                            registro_anterior = Registro_Auditorias_HEB.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Agosto',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Octubre':
                            registro_anterior = Registro_Auditorias_HEB.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=30, 
                                MES='Septiembre',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Noviembre':
                            registro_anterior = Registro_Auditorias_HEB.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Octubre',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Diciembre':
                            registro_anterior = Registro_Auditorias_HEB.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=30, 
                                MES='Noviembre',
                                AÑO=año_seleccionado
                            ).first()
                        if registro_anterior:
                            producto.DIA_ANTERIOR = registro_anterior.DIA_ACTUAL
                            producto.DIA_ACTUAL = 0.0
                            producto.CAJAS = 0
                            producto.COMPRAS = 0.0
                            producto.ENTRADA = 0.0
                            producto.SALIDA = 0.0
                            producto.MERMAS = 0.0
                            producto.save()
                    # Si no existe un registro, busca el registro del día anterior (DIA - 1)
                    else: 
                        registro_anterior = Registro_Auditorias_HEB.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=dia_seleccionado - 1,  # Restar 1 al día
                            MES=mes_seleccionado,
                            AÑO=año_seleccionado
                            ).first()
                        if registro_anterior:
                            producto.DIA_ANTERIOR = registro_anterior.DIA_ACTUAL
                            producto.DIA_ACTUAL = 0.0
                            producto.CAJAS = 0
                            producto.COMPRAS = 0.0
                            producto.ENTRADA = 0.0
                            producto.SALIDA = 0.0
                            producto.MERMAS = 0.0

        return render(request, 'cuentas/auditoria.html', {
            'productos': productos,
            'dias': dias,
            'años': años,
            'dia_semanal_seleccionado': dia_semanal_seleccionado,
            'dia_seleccionado': dia_seleccionado,
            'mes_seleccionado': mes_seleccionado,
            'año_seleccionado': año_seleccionado,
        })

    # Redirigir de nuevo a la página de auditoria después de procesar los datos
    return redirect('cuentas:auditoria')


from collections import defaultdict

def resumenHEB(request):
    # Generar una lista de días del 1 al 31
    dias = list(range(1, 32))
    # Generar una lista de años desde 2020 hasta 2050
    años = list(range(2023, 2051))

    if request.method == 'POST':
        mes_seleccionado = request.POST.get('mes', '')
        año_seleccionado = int(request.POST.get('anio', 0))
        dia_inicial = int(request.POST.get('dia_inicial', 0))
        dia_final = int(request.POST.get('dia_final', 0))
        
        registros = Registro_Auditorias_HEB.objects.filter(MES=mes_seleccionado, AÑO=año_seleccionado)
        registros2 = registros.filter(DIA__gte=dia_inicial, DIA__lte=dia_final)

        # Crear un diccionario para almacenar los registros por DESCRIPCION
        registros_por_descripcion = defaultdict(list)
        for registro in registros:
            descripcion = registro.DESCRIPCION
            registros_por_descripcion[descripcion].append(registro)

        registros_agrupados = []
        for descripcion, registros in registros_por_descripcion.items():
            registro_agrupado = {
                'descripcion': descripcion,
                'medida': registros[0].MEDIDA,  # Tomamos la MEDIDA del primer registro
                'consumos_diarios': [registro.CONSUMO_DIARIO for registro in registros]
            }
            registros_agrupados.append(registro_agrupado)
        
        registros_por_dia = {}
        for registro in registros:
            dia = registro.DIA
            if dia not in registros_por_dia:
                registros_por_dia[dia] = []
                registros_por_dia[dia].append(registro)

        registros_agrupados2 = []  # Inicializa la lista fuera del bucle
        for descripcion, registros in registros_por_descripcion.items():
            consumos_diarios = [registro.CONSUMO_DIARIO for registro in registros if registro in registros2]
            promedio = sum(consumos_diarios) / len(consumos_diarios) if consumos_diarios else 0
            suma = sum(consumos_diarios)
            promedio_redondeado = round(promedio, 3)
            suma_redondeada = round(suma,3)
            registro_agrupado = {
                'descripcion': descripcion,
                'medida': registros[0].MEDIDA,
                'consumos_diarios': consumos_diarios,
                'promedio': promedio_redondeado,
                'suma': suma_redondeada,
            }
            registros_agrupados2.append(registro_agrupado)

        return render(request, 'cuentas/resumen_HEB.html', {
            'dias': dias,
            'años': años,
            'registros_agrupados': registros_agrupados,
            'mes_seleccionado': mes_seleccionado,
            'año_seleccionado': año_seleccionado,
            'registros_por_dia': registros_por_dia,
            'registros_agrupados2': registros_agrupados2,
            'dia_inicial': dia_inicial,  # Pasar los valores de día inicial y final
            'dia_final': dia_final,
        })

    return render(request, 'cuentas/resumen_HEB.html', {'dias':dias,'años': años})

def auditoriaBebidas(request):
    productos = Auditorias_HEB_Bebidas.objects.all()

    # Generar una lista de días del 1 al 31
    dias = list(range(1, 32))

    # Generar una lista de años desde 2020 hasta 2050
    años = list(range(2023, 2051))

    if request.method == 'POST' and 'limpiar_Bebidas_HEB' in request.POST:

        for producto in productos:
            producto.DIA = 0
            producto.MES = 'Mes'
            producto.AÑO = 0 
            producto.DIA_SEMANAL = 'Dia'
            producto.DIA_ANTERIOR = 0
            producto.DIA_ACTUAL = 0
            producto.COMPRAS = 0
            producto.ENTRADA = 0
            producto.SALIDA = 0
            producto.CONSUMO_DIARIO = 0
            producto.save()
        return render(request, 'cuentas/auditoriaBebidas.html', {'productos': productos, 'dias': dias, 'años': años})
    
    if request.method == 'POST':
        if 'archivo' in request.FILES:
            archivo = request.FILES['archivo']
            workbook = openpyxl.load_workbook(archivo)
            sheet = workbook.active

            productos_nuevos = []

            for row in sheet.iter_rows(min_row=4, values_only=True):
                DESCRIPCION, MEDIDA = row[1], str(row[2])

                if DESCRIPCION and MEDIDA:
                    producto, created = Auditorias_HEB_Bebidas.objects.get_or_create(
                        DESCRIPCION=DESCRIPCION,
                        defaults={
                            'MEDIDA': MEDIDA,
                            'DIA_ANTERIOR': 0,
                            'DIA_ACTUAL': 0,
                            'COMPRAS': 0,
                            'ENTRADA': 0,
                            'SALIDA': 0,
                            'CONSUMO_DIARIO': 0,
                            'DIA_SEMANAL': 'Dia',
                            'DIA': 0,
                            'MES': 'Mes',
                            'AÑO': 0,
                        }
                    )
                    if created:
                        productos_nuevos.append(producto)
                    else:
                        producto.DESCRIPCION = DESCRIPCION
                        producto.MEDIDA = MEDIDA
                        producto.save()

    if request.method == 'POST' and 'guardar_cambios' in request.POST:
        # Procesar la selección de DÍA, MES y AÑO
        dia_seleccionado = int(request.POST.get('dia', 0))
        mes_seleccionado = request.POST.get('mes', '')
        año_seleccionado = int(request.POST.get('anio', 0))
        dia_semanal_seleccionado = request.POST.get('dia_semanal', '')

        # Iterar sobre los productos y actualizar los campos correspondientes
        for producto in productos:
            dia_anterior = int(request.POST.get(f'dia_anterior_{producto.id}', 0))
            dia_actual = int(request.POST.get(f'dia_actual_{producto.id}', 0))
            compras = int(request.POST.get(f'compras_{producto.id}', 0))
            entrada = int(request.POST.get(f'entrada_{producto.id}', 0))
            salida = int(request.POST.get(f'salida_{producto.id}', 0))

            # Actualizar los campos del producto
            producto.DIA = dia_seleccionado
            producto.MES = mes_seleccionado
            producto.AÑO = año_seleccionado
            producto.DIA_SEMANAL = dia_semanal_seleccionado
            producto.DIA_ANTERIOR = dia_anterior
            producto.DIA_ACTUAL = dia_actual
            producto.COMPRAS = compras
            producto.ENTRADA = entrada
            producto.SALIDA = salida

            # Calcular el consumo diario
            if dia_actual == 0:
                producto.CONSUMO_DIARIO = 0
            else:
                producto.CONSUMO_DIARIO = (-dia_anterior) - compras + dia_actual - entrada + salida 

            producto.save()

                # Verificar si existe un registro en Registro_Auditorias_HEB con la misma fecha
            registro_existente = Registro_Auditorias_HEB_Bebidas.objects.filter(
                DESCRIPCION=producto.DESCRIPCION,
                MEDIDA=producto.MEDIDA,
                DIA=dia_seleccionado,
                MES=mes_seleccionado,
                AÑO=año_seleccionado
            ).first()

            if registro_existente:
                 # Si existe un registro con la misma fecha, actualizar los datos en Registro_Auditorias_HEB
                registro_existente.DIA_ANTERIOR = producto.DIA_ANTERIOR
                registro_existente.DIA_ACTUAL = producto.DIA_ACTUAL
                registro_existente.COMPRAS = producto.COMPRAS
                registro_existente.ENTRADA = producto.ENTRADA
                registro_existente.SALIDA = producto.SALIDA
                registro_existente.CONSUMO_DIARIO = producto.CONSUMO_DIARIO
                registro_existente.save()
            else:
            # Si no existe un registro con la misma fecha, insertar un nuevo registro en Registro_Auditorias_HEB
                nuevo_registro = Registro_Auditorias_HEB_Bebidas(
                    DESCRIPCION=producto.DESCRIPCION,
                    MEDIDA=producto.MEDIDA,
                    DIA_ANTERIOR=producto.DIA_ANTERIOR,
                    DIA_ACTUAL=producto.DIA_ACTUAL,
                    COMPRAS=producto.COMPRAS,
                    ENTRADA=producto.ENTRADA,
                    SALIDA=producto.SALIDA,
                    CONSUMO_DIARIO=producto.CONSUMO_DIARIO,
                    DIA_SEMANAL = dia_semanal_seleccionado,
                    DIA=dia_seleccionado,
                    MES=mes_seleccionado,
                    AÑO=año_seleccionado
                    )
                nuevo_registro.save()
                
        return render(request, 'cuentas/auditoriaBebidas.html', {
            'dias': dias,
            'años': años,
            'productos': productos,
            'dia_semanal_seleccionado': dia_semanal_seleccionado,
            'dia_seleccionado': dia_seleccionado,
            'mes_seleccionado': mes_seleccionado,
            'año_seleccionado': año_seleccionado,
        })

    return render(request, 'cuentas/auditoriaBebidas.html', {'productos': productos, 'dias': dias, 'años': años})

def recuperar_datos_Bebidas(request):
    # Generar una lista de días del 1 al 31
    dias = list(range(1, 32))

    # Generar una lista de años desde 2020 hasta 2050
    años = list(range(2023, 2051))

    if request.method == 'POST':
        # Obtener el día, mes y año seleccionados
        dia_seleccionado = int(request.POST.get('dia', 0))
        mes_seleccionado = request.POST.get('mes', '')
        año_seleccionado = int(request.POST.get('anio', 0))
        dia_semanal_seleccionado = request.POST.get('dia_semanal', '')
        
        productos = Auditorias_HEB_Bebidas.objects.all()
        
        for producto in productos:
            registro = Registro_Auditorias_HEB_Bebidas.objects.filter(
                DESCRIPCION=producto.DESCRIPCION,
                MEDIDA=producto.MEDIDA,
                DIA=dia_seleccionado,
                MES=mes_seleccionado,
                AÑO=año_seleccionado
            ).first()
            
            if registro:
                # Si existen registros, actualizar los datos en Auditorias_HEB
                producto.DIA_ANTERIOR = registro.DIA_ANTERIOR
                producto.DIA_ACTUAL = registro.DIA_ACTUAL
                producto.COMPRAS = registro.COMPRAS
                producto.ENTRADA = registro.ENTRADA
                producto.SALIDA = registro.SALIDA
                producto.CONSUMO_DIARIO = registro.CONSUMO_DIARIO
                producto.DIA_SEMANAL = registro.DIA_SEMANAL
                producto.DIA = registro.DIA
                producto.MES = registro.MES
                producto.AÑO = registro.AÑO
                producto.save()
            else:
                if dia_seleccionado == 1 and año_seleccionado == 2023:
                    if mes_seleccionado == 'Septiembre':
                        registro_anterior = Registro_Auditorias_HEB_Bebidas.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=31, 
                            MES='Agosto',
                            AÑO=año_seleccionado
                            ).first()
                    elif mes_seleccionado == 'Octubre':
                        registro_anterior = Registro_Auditorias_HEB_Bebidas.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=30, 
                            MES='Septiembre',
                            AÑO=año_seleccionado
                            ).first()
                    elif mes_seleccionado == 'Noviembre':
                        registro_anterior = Registro_Auditorias_HEB_Bebidas.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=31, 
                            MES='Octubre',
                            AÑO=año_seleccionado
                            ).first()
                    elif mes_seleccionado == 'Diciembre':
                        registro_anterior = Registro_Auditorias_HEB_Bebidas.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=30, 
                            MES='Noviembre',
                            AÑO=año_seleccionado
                            ).first()
                    if registro_anterior:
                        producto.DIA_ANTERIOR = registro_anterior.DIA_ACTUAL
                        producto.DIA_ACTUAL = 0
                        producto.COMPRAS = 0
                        producto.ENTRADA = 0
                        producto.SALIDA = 0
                        producto.save()
                else:
                    if dia_seleccionado == 1 and año_seleccionado == 2024:
                        if mes_seleccionado == 'Enero':
                            registro_anterior = Registro_Auditorias_HEB_Bebidas.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Diciembre',
                                AÑO=2023
                                ).first()
                        if mes_seleccionado == 'Febrero':
                            registro_anterior = Registro_Auditorias_HEB_Bebidas.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Enero',
                                AÑO=año_seleccionado
                                ).first()
                        elif mes_seleccionado == 'Marzo':
                            registro_anterior = Registro_Auditorias_HEB_Bebidas.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=28, 
                                MES='Febrero',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Abril':
                            registro_anterior = Registro_Auditorias_HEB_Bebidas.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=31, 
                            MES='Marzo',
                            AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Mayo':
                            registro_anterior = Registro_Auditorias_HEB_Bebidas.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=30, 
                                MES='Abril',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Junio':
                            registro_anterior = Registro_Auditorias_HEB_Bebidas.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Mayo',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Julio':
                            registro_anterior = Registro_Auditorias_HEB_Bebidas.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=30, 
                                MES='Junio',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Agosto':
                            registro_anterior = Registro_Auditorias_HEB_Bebidas.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Julio',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Septiembre':
                            registro_anterior = Registro_Auditorias_HEB_Bebidas.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Agosto',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Octubre':
                            registro_anterior = Registro_Auditorias_HEB_Bebidas.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=30, 
                                MES='Septiembre',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Noviembre':
                            registro_anterior = Registro_Auditorias_HEB_Bebidas.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Octubre',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Diciembre':
                            registro_anterior = Registro_Auditorias_HEB_Bebidas.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=30, 
                                MES='Noviembre',
                                AÑO=año_seleccionado
                            ).first()
                        if registro_anterior:
                            producto.DIA_ANTERIOR = registro_anterior.DIA_ACTUAL
                            producto.DIA_ACTUAL = 0
                            producto.COMPRAS = 0
                            producto.ENTRADA = 0
                            producto.SALIDA = 0
                            producto.save()
                    # Si no existe un registro, busca el registro del día anterior (DIA - 1)
                    else: 
                        registro_anterior = Registro_Auditorias_HEB_Bebidas.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=dia_seleccionado - 1,  # Restar 1 al día
                            MES=mes_seleccionado,
                            AÑO=año_seleccionado
                            ).first()
                        if registro_anterior:
                            producto.DIA_ANTERIOR = registro_anterior.DIA_ACTUAL
                            producto.DIA_ACTUAL = 0
                            producto.COMPRAS = 0
                            producto.ENTRADA = 0
                            producto.SALIDA = 0
                            producto.save()

        return render(request, 'cuentas/auditoriaBebidas.html', {
            'productos': productos,
            'dias': dias,
            'años': años,
            'dia_semanal_seleccionado': dia_semanal_seleccionado,
            'dia_seleccionado': dia_seleccionado,
            'mes_seleccionado': mes_seleccionado,
            'año_seleccionado': año_seleccionado,
        })

    # Redirigir de nuevo a la página de auditoria después de procesar los datos
    return redirect('cuentas:auditoriaBebidas')

def resumenHEB_Bebidas(request):
    # Generar una lista de días del 1 al 31
    dias = list(range(1, 32))
    # Generar una lista de años desde 2020 hasta 2050
    años = list(range(2023, 2051))

    if request.method == 'POST':
        mes_seleccionado = request.POST.get('mes', '')
        año_seleccionado = int(request.POST.get('anio', 0))
        dia_inicial = int(request.POST.get('dia_inicial', 0))
        dia_final = int(request.POST.get('dia_final', 0))
        
        registros = Registro_Auditorias_HEB_Bebidas.objects.filter(MES=mes_seleccionado, AÑO=año_seleccionado)
        registros2 = registros.filter(DIA__gte=dia_inicial, DIA__lte=dia_final)

        # Crear un diccionario para almacenar los registros por DESCRIPCION
        registros_por_descripcion = defaultdict(list)
        for registro in registros:
            descripcion = registro.DESCRIPCION
            registros_por_descripcion[descripcion].append(registro)

        registros_agrupados = []
        for descripcion, registros in registros_por_descripcion.items():
            registro_agrupado = {
                'descripcion': descripcion,
                'medida': registros[0].MEDIDA,  # Tomamos la MEDIDA del primer registro
                'consumos_diarios': [registro.CONSUMO_DIARIO for registro in registros]
            }
            registros_agrupados.append(registro_agrupado)
        
        registros_por_dia = {}
        for registro in registros:
            dia = registro.DIA
            if dia not in registros_por_dia:
                registros_por_dia[dia] = []
                registros_por_dia[dia].append(registro)

        registros_agrupados2 = []  # Inicializa la lista fuera del bucle
        for descripcion, registros in registros_por_descripcion.items():
            consumos_diarios = [registro.CONSUMO_DIARIO for registro in registros if registro in registros2]
            promedio = sum(consumos_diarios) / len(consumos_diarios) if consumos_diarios else 0
            suma = sum(consumos_diarios)
            promedio_redondeado = round(promedio, 3)
            suma_redondeada = round(suma,3)
            registro_agrupado = {
                'descripcion': descripcion,
                'medida': registros[0].MEDIDA,
                'consumos_diarios': consumos_diarios,
                'promedio': promedio_redondeado,
                'suma': suma_redondeada,
            }
            registros_agrupados2.append(registro_agrupado)
            

        return render(request, 'cuentas/resumen_HEB_Bebidas.html', {
            'dias': dias,
            'años': años,
            'registros_agrupados': registros_agrupados,
            'mes_seleccionado': mes_seleccionado,
            'año_seleccionado': año_seleccionado,
            'registros_por_dia': registros_por_dia,
            'registros_agrupados2': registros_agrupados2,
            'dia_inicial': dia_inicial,  # Pasar los valores de día inicial y final
            'dia_final': dia_final,
        })

    return render(request, 'cuentas/resumen_HEB_Bebidas.html', {'dias':dias,'años': años})

def auditoria_4C(request):
    productos = Auditorias_4C.objects.all()

    # Generar una lista de días del 1 al 31
    dias = list(range(1, 32))

    # Generar una lista de años desde 2020 hasta 2050
    años = list(range(2023, 2051))
    if request.method == 'POST' and 'limpiar_4C' in request.POST:

        for producto in productos:
            producto.DIA = 0
            producto.MES = 'Mes'
            producto.AÑO = 0 
            producto.DIA_SEMANAL = 'Dia'
            producto.DIA_ANTERIOR = 0.0
            producto.DIA_ACTUAL = 0.0
            producto.CAJAS = 0
            producto.COMPRAS = 0.0
            producto.ENTRADA = 0.0
            producto.SALIDA = 0.0
            producto.MERMAS = 0.0
            producto.CONSUMO_DIARIO = 0.0
            producto.save()
        return render(request, 'cuentas/auditoria_4C.html', {'productos': productos, 'dias': dias, 'años': años})

    if request.method == 'POST':
        if 'archivo' in request.FILES:
            archivo = request.FILES['archivo']
            workbook = openpyxl.load_workbook(archivo)
            sheet = workbook.active

            productos_nuevos = []

            for row in sheet.iter_rows(min_row=4, values_only=True):
                DESCRIPCION, MEDIDA = row[1], str(row[2])

                if DESCRIPCION and MEDIDA:
                    producto, created = Auditorias_4C.objects.get_or_create(
                        DESCRIPCION=DESCRIPCION,
                        defaults={
                            'MEDIDA': MEDIDA,
                            'DIA_ANTERIOR': 0.0,
                            'DIA_ACTUAL': 0.0,
                            'CAJAS': 0,
                            'COMPRAS': 0.0,
                            'ENTRADA': 0.0,
                            'SALIDA': 0.0,
                            'MERMAS': 0.0,
                            'CONSUMO_DIARIO': 0.0,
                            'DIA_SEMANAL': 'Dia',
                            'DIA': 0,
                            'MES': 'Mes',
                            'AÑO': 0,
                        }
                    )
                    if created:
                        productos_nuevos.append(producto)
                    else:
                        producto.DESCRIPCION = DESCRIPCION
                        producto.MEDIDA = MEDIDA
                        producto.save()
    
    if request.method == 'POST' and 'guardar_cambios' in request.POST:
        # Procesar la selección de DÍA, MES y AÑO
        dia_seleccionado = int(request.POST.get('dia', 0))
        mes_seleccionado = request.POST.get('mes', '')
        año_seleccionado = int(request.POST.get('anio', 0))
        dia_semanal_seleccionado = request.POST.get('dia_semanal', '')
        

        # Iterar sobre los productos y actualizar los campos correspondientes
        for producto in productos:
            dia_anterior = float(request.POST.get(f'dia_anterior_{producto.id}', 0))
            dia_actual = float(request.POST.get(f'dia_actual_{producto.id}', 0))
            cajas = int(request.POST.get(f'cajas_{producto.id}', 0))
            compras = float(request.POST.get(f'compras_{producto.id}', 0))
            entrada = float(request.POST.get(f'entrada_{producto.id}', 0))
            salida = float(request.POST.get(f'salida_{producto.id}', 0))
            mermas = float(request.POST.get(f'mermas_{producto.id}', 0))

            # Actualizar los campos del producto
            producto.DIA = dia_seleccionado
            producto.MES = mes_seleccionado
            producto.AÑO = año_seleccionado
            producto.DIA_SEMANAL = dia_semanal_seleccionado
            producto.DIA_ANTERIOR = dia_anterior
            producto.DIA_ACTUAL = dia_actual
            producto.CAJAS = cajas
            if producto.id == 1:
                producto.COMPRAS = producto.CAJAS * 70
            elif producto.id == 2:
                producto.COMPRAS = producto.CAJAS * 9.6
            elif producto.id == 3:
                producto.COMPRAS = producto.CAJAS * 11.5
            elif producto.id == 4:
                producto.COMPRAS = producto.CAJAS * 8
            elif producto.id == 5:
                producto.COMPRAS = producto.CAJAS * 7
            elif producto.id == 6:
                producto.COMPRAS = producto.CAJAS * 10
            elif producto.id == 7:
                producto.COMPRAS = producto.CAJAS * 9.6
            elif producto.id == 8:
                producto.COMPRAS = producto.CAJAS * 6
            elif producto.id == 9:
                producto.COMPRAS = 0.0
            elif producto.id == 10:
                producto.COMPRAS = producto.CAJAS * 12
            elif producto.id == 11:
                producto.COMPRAS = producto.CAJAS * 75
            elif producto.id == 12:
                producto.COMPRAS = producto.CAJAS * 600
            elif producto.id == 13:
                producto.COMPRAS = producto.CAJAS * 7
            elif producto.id == 14:
                producto.COMPRAS = producto.CAJAS * 9.08
            elif producto.id == 15:
                producto.COMPRAS = producto.CAJAS * 9.08
            elif producto.id == 16:
                producto.COMPRAS = producto.CAJAS * 9
            elif producto.id == 17:
                producto.COMPRAS = producto.CAJAS * 8
            elif producto.id == 18:
                producto.COMPRAS = producto.CAJAS * 8
            elif producto.id == 19:
                producto.COMPRAS = producto.CAJAS * 8
            elif producto.id == 20:
                producto.COMPRAS = producto.CAJAS * 8
            elif producto.id == 21:
                producto.COMPRAS = producto.CAJAS * 8
            elif producto.id == 22:
                producto.COMPRAS = producto.CAJAS * 8
            elif producto.id == 23:
                producto.COMPRAS = producto.CAJAS * 4
            elif producto.id == 24:
                producto.COMPRAS = producto.CAJAS * 8
            elif producto.id == 25:
                producto.COMPRAS = producto.CAJAS * 4
            elif producto.id == 26:
                producto.COMPRAS = producto.CAJAS * 6
            elif producto.id == 27:
                producto.COMPRAS = producto.CAJAS * 6
            elif producto.id == 28:
                producto.COMPRAS = producto.CAJAS * 50
            elif producto.id == 29:
                producto.COMPRAS = producto.CAJAS * 200
            elif producto.id == 30:
                producto.COMPRAS = producto.CAJAS
            elif producto.id == 31:
                producto.COMPRAS = 0.0
            producto.ENTRADA = entrada
            producto.SALIDA = salida
            producto.MERMAS = mermas

            # Calcular el consumo diario
            if dia_actual == 0:
                producto.CONSUMO_DIARIO = 0
            else:
                producto.CONSUMO_DIARIO = (-dia_anterior) - producto.COMPRAS + dia_actual - entrada + salida - mermas

            producto.save()

            if dia_semanal_seleccionado == 'Viernes':
                    # Verificar si existe un registro en Registro_Auditorias_4C con la misma fecha
                registro_existente = Registro_Auditorias_4C.objects.filter(
                    DESCRIPCION=producto.DESCRIPCION,
                    MEDIDA=producto.MEDIDA,
                    DIA_SEMANAL = dia_semanal_seleccionado,
                    DIA=dia_seleccionado,
                    MES=mes_seleccionado,
                    AÑO=año_seleccionado
                    ).first()
                
                if dia_seleccionado == 1:
                    if mes_seleccionado == 'Septiembre' and año_seleccionado == 2023:
                        registro_anterior = Registro_Auditorias_4C.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=31, 
                        MES='Agosto',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Octubre' and año_seleccionado == 2023:
                        registro_anterior = Registro_Auditorias_4C.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=30, 
                        MES='Septiembre',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Noviembre' and año_seleccionado == 2023:
                        registro_anterior = Registro_Auditorias_4C.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=31, 
                        MES='Octubre',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Diciembre' and año_seleccionado == 2023:
                        registro_anterior = Registro_Auditorias_4C.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=30, 
                        MES='Noviembre',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Enero' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_4C.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=31, 
                        MES='Diciembre',
                        AÑO=2023
                        ).first()
                    elif mes_seleccionado == 'Febrero' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_4C.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=31, 
                        MES='Enero',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Marzo' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_4C.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=28, 
                        MES='Febrero',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Abril' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_4C.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=31, 
                        MES='Marzo',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Mayo' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_4C.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=30, 
                        MES='Abril',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Junio' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_4C.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=31, 
                        MES='Mayo',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Julio' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_4C.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=30, 
                        MES='Junio',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Agosto' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_4C.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=31, 
                        MES='Julio',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Septiembre' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_4C.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=31, 
                        MES='Agosto',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Octubre' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_4C.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=30, 
                        MES='Septiembre',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Noviembre' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_4C.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=31, 
                        MES='Octubre',
                        AÑO=año_seleccionado
                        ).first()
                    elif mes_seleccionado == 'Diciembre' and año_seleccionado == 2024:
                        registro_anterior = Registro_Auditorias_4C.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=30, 
                        MES='Noviembre',
                        AÑO=año_seleccionado
                        ).first()
                else:
                    registro_anterior = Registro_Auditorias_4C.objects.filter(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA=dia_seleccionado - 1, 
                        MES=mes_seleccionado,
                        AÑO=año_seleccionado
                        ).first()

                if registro_existente:
                    checkbox_name = f'check_{producto.id}'  # Nombre del campo checkbox
                    if request.POST.get(checkbox_name) == 'on':
                        producto.DIA_ANTERIOR = round(producto.DIA_ANTERIOR - producto.COMPRAS, 3)
                        producto.CONSUMO_DIARIO = (-producto.DIA_ANTERIOR) - producto.COMPRAS + dia_actual - entrada + salida - mermas
                        producto.save()
                        registro_anterior.DIA_ACTUAL = producto.DIA_ANTERIOR
                        registro_anterior.CONSUMO_DIARIO = (-registro_anterior.DIA_ANTERIOR) - registro_anterior.COMPRAS + registro_anterior.DIA_ACTUAL - registro_anterior.ENTRADA + registro_anterior.SALIDA - registro_anterior.MERMAS
                        registro_anterior.save()
                    else:
                        if registro_anterior:
                            registro_anterior.DIA_ACTUAL = producto.DIA_ANTERIOR
                            registro_anterior.CONSUMO_DIARIO = (-registro_anterior.DIA_ANTERIOR) - registro_anterior.COMPRAS + registro_anterior.DIA_ACTUAL - registro_anterior.ENTRADA + registro_anterior.SALIDA - registro_anterior.MERMAS
                            registro_anterior.save()
                 # Si existe un registro con la misma fecha, actualizar los datos en Registro_Auditorias_HEB
                    registro_existente.DIA_ANTERIOR = producto.DIA_ANTERIOR 
                    registro_existente.DIA_ACTUAL = producto.DIA_ACTUAL
                    registro_existente.CAJAS = producto.CAJAS
                    registro_existente.COMPRAS = producto.COMPRAS
                    registro_existente.ENTRADA = producto.ENTRADA
                    registro_existente.SALIDA = producto.SALIDA
                    registro_existente.MERMAS = producto.MERMAS
                    registro_existente.CONSUMO_DIARIO = producto.CONSUMO_DIARIO
                    registro_existente.save()
                else:
            # Si no existe un registro con la misma fecha, insertar un nuevo registro en Registro_Auditorias_4C
                    checkbox_name = f'check_{producto.id}'  # Nombre del campo checkbox
                    if request.POST.get(checkbox_name) == 'on':
                        producto.DIA_ANTERIOR = round(producto.DIA_ANTERIOR - producto.COMPRAS, 3)
                        producto.CONSUMO_DIARIO = (-producto.DIA_ANTERIOR) - producto.COMPRAS + dia_actual - entrada + salida - mermas
                        producto.save()
                        registro_anterior.DIA_ACTUAL = producto.DIA_ANTERIOR
                        registro_anterior.CONSUMO_DIARIO = (-registro_anterior.DIA_ANTERIOR) - registro_anterior.COMPRAS + registro_anterior.DIA_ACTUAL - registro_anterior.ENTRADA + registro_anterior.SALIDA - registro_anterior.MERMAS
                        registro_anterior.save()
                    else:
                        if registro_anterior:
                            registro_anterior.DIA_ACTUAL = producto.DIA_ANTERIOR
                            registro_anterior.CONSUMO_DIARIO = (-registro_anterior.DIA_ANTERIOR) - registro_anterior.COMPRAS + registro_anterior.DIA_ACTUAL - registro_anterior.ENTRADA + registro_anterior.SALIDA - registro_anterior.MERMAS
                            registro_anterior.save()
                    nuevo_registro = Registro_Auditorias_4C(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA_ANTERIOR=producto.DIA_ANTERIOR,
                        DIA_ACTUAL=producto.DIA_ACTUAL,
                        CAJAS=producto.CAJAS,
                        COMPRAS=producto.COMPRAS,
                        ENTRADA=producto.ENTRADA,
                        SALIDA=producto.SALIDA,
                        MERMAS=producto.MERMAS,
                        CONSUMO_DIARIO=producto.CONSUMO_DIARIO,
                        DIA_SEMANAL = dia_semanal_seleccionado,
                        DIA=dia_seleccionado,
                        MES=mes_seleccionado,
                        AÑO=año_seleccionado
                    )
                    nuevo_registro.save()
                    
            else:
                    # Verificar si existe un registro en Registro_Auditorias_4C con la misma fecha
                registro_existente = Registro_Auditorias_4C.objects.filter(
                    DESCRIPCION=producto.DESCRIPCION,
                    MEDIDA=producto.MEDIDA,
                    DIA_SEMANAL = dia_semanal_seleccionado,
                    DIA=dia_seleccionado,
                    MES=mes_seleccionado,
                    AÑO=año_seleccionado
                    ).first()

                if registro_existente:
                 # Si existe un registro con la misma fecha, actualizar los datos en Registro_Auditorias_4C
                    registro_existente.DIA_ANTERIOR = producto.DIA_ANTERIOR
                    registro_existente.DIA_ACTUAL = producto.DIA_ACTUAL
                    registro_existente.CAJAS = producto.CAJAS
                    registro_existente.COMPRAS = producto.COMPRAS
                    registro_existente.ENTRADA = producto.ENTRADA
                    registro_existente.SALIDA = producto.SALIDA
                    registro_existente.MERMAS = producto.MERMAS
                    registro_existente.CONSUMO_DIARIO = producto.CONSUMO_DIARIO
                    registro_existente.save()
                else:
            # Si no existe un registro con la misma fecha, insertar un nuevo registro en Registro_Auditorias_4C
                    nuevo_registro = Registro_Auditorias_4C(
                        DESCRIPCION=producto.DESCRIPCION,
                        MEDIDA=producto.MEDIDA,
                        DIA_ANTERIOR=producto.DIA_ANTERIOR,
                        DIA_ACTUAL=producto.DIA_ACTUAL,
                        CAJAS=producto.CAJAS,
                        COMPRAS=producto.COMPRAS,
                        ENTRADA=producto.ENTRADA,
                        SALIDA=producto.SALIDA,
                        MERMAS=producto.MERMAS,
                        CONSUMO_DIARIO=producto.CONSUMO_DIARIO,
                        DIA_SEMANAL = dia_semanal_seleccionado,
                        DIA=dia_seleccionado,
                        MES=mes_seleccionado,
                        AÑO=año_seleccionado
                        )
                    nuevo_registro.save()
            
                
        return render(request, 'cuentas/auditoria_4C.html', {
            'dias': dias,
            'años': años,
            'productos': productos,
            'dia_semanal_seleccionado': dia_semanal_seleccionado,
            'dia_seleccionado': dia_seleccionado,
            'mes_seleccionado': mes_seleccionado,
            'año_seleccionado': año_seleccionado,
        })

    return render(request, 'cuentas/auditoria_4C.html', {'productos': productos, 'dias': dias, 'años': años})

def recuperar_datos_4C(request):
      # Generar una lista de días del 1 al 31
    dias = list(range(1, 32))

    # Generar una lista de años desde 2020 hasta 2050
    años = list(range(2023, 2051))
    if request.method == 'POST':
        # Obtener el día, mes y año seleccionados
        dia_seleccionado = int(request.POST.get('dia', 0))
        mes_seleccionado = request.POST.get('mes', '')
        año_seleccionado = int(request.POST.get('anio', 0))
        dia_semanal_seleccionado = request.POST.get('dia_semanal', '')
        productos = Auditorias_4C.objects.all()
        
        for producto in productos:
            registro = Registro_Auditorias_4C.objects.filter(
                DESCRIPCION=producto.DESCRIPCION,
                MEDIDA=producto.MEDIDA,
                DIA=dia_seleccionado,
                MES=mes_seleccionado,
                AÑO=año_seleccionado
            ).first()
            
            if registro:
                # Si existen registros, actualizar los datos en Auditorias_4C
                producto.DIA_ANTERIOR = registro.DIA_ANTERIOR
                producto.DIA_ACTUAL = registro.DIA_ACTUAL
                producto.CAJAS = registro.CAJAS
                producto.COMPRAS = registro.COMPRAS
                producto.ENTRADA = registro.ENTRADA
                producto.SALIDA = registro.SALIDA
                producto.MERMAS = registro.MERMAS
                producto.CONSUMO_DIARIO = registro.CONSUMO_DIARIO
                producto.DIA_SEMANAL = registro.DIA_SEMANAL
                producto.DIA = registro.DIA
                producto.MES = registro.MES
                producto.AÑO = registro.AÑO
                producto.save()
            else:
                if dia_seleccionado == 1 and año_seleccionado == 2023:
                    if mes_seleccionado == 'Septiembre':
                        registro_anterior = Registro_Auditorias_4C.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=31, 
                            MES='Agosto',
                            AÑO=año_seleccionado
                            ).first()
                    elif mes_seleccionado == 'Octubre':
                        registro_anterior = Registro_Auditorias_4C.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=30, 
                            MES='Septiembre',
                            AÑO=año_seleccionado
                            ).first()
                    elif mes_seleccionado == 'Noviembre':
                        registro_anterior = Registro_Auditorias_4C.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=31, 
                            MES='Octubre',
                            AÑO=año_seleccionado
                            ).first()
                    elif mes_seleccionado == 'Diciembre':
                        registro_anterior = Registro_Auditorias_4C.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=30, 
                            MES='Noviembre',
                            AÑO=año_seleccionado
                            ).first()
                    if registro_anterior:
                        producto.DIA_ANTERIOR = registro_anterior.DIA_ACTUAL
                        producto.DIA_ACTUAL = 0.0
                        producto.CAJAS = 0
                        producto.COMPRAS = 0.0
                        producto.ENTRADA = 0.0
                        producto.SALIDA = 0.0
                        producto.MERMAS = 0.0
                        producto.save()
                else:
                    if dia_seleccionado == 1 and año_seleccionado == 2024:
                        if mes_seleccionado == 'Enero':
                            registro_anterior = Registro_Auditorias_4C.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Diciembre',
                                AÑO=2023
                                ).first()
                        if mes_seleccionado == 'Febrero':
                            registro_anterior = Registro_Auditorias_4C.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Enero',
                                AÑO=año_seleccionado
                                ).first()
                        elif mes_seleccionado == 'Marzo':
                            registro_anterior = Registro_Auditorias_4C.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=28, 
                                MES='Febrero',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Abril':
                            registro_anterior = Registro_Auditorias_4C.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=31, 
                            MES='Marzo',
                            AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Mayo':
                            registro_anterior = Registro_Auditorias_4C.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=30, 
                                MES='Abril',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Junio':
                            registro_anterior = Registro_Auditorias_4C.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Mayo',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Julio':
                            registro_anterior = Registro_Auditorias_4C.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=30, 
                                MES='Junio',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Agosto':
                            registro_anterior = Registro_Auditorias_4C.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Julio',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Septiembre':
                            registro_anterior = Registro_Auditorias_4C.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Agosto',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Octubre':
                            registro_anterior = Registro_Auditorias_4C.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=30, 
                                MES='Septiembre',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Noviembre':
                            registro_anterior = Registro_Auditorias_4C.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Octubre',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Diciembre':
                            registro_anterior = Registro_Auditorias_4C.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=30, 
                                MES='Noviembre',
                                AÑO=año_seleccionado
                            ).first()
                        if registro_anterior:
                            producto.DIA_ANTERIOR = registro_anterior.DIA_ACTUAL
                            producto.DIA_ACTUAL = 0.0
                            producto.CAJAS = 0
                            producto.COMPRAS = 0.0
                            producto.ENTRADA = 0.0
                            producto.SALIDA = 0.0
                            producto.MERMAS = 0.0
                            producto.save()
                    # Si no existe un registro, busca el registro del día anterior (DIA - 1)
                    else: 
                        registro_anterior = Registro_Auditorias_4C.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=dia_seleccionado - 1,  # Restar 1 al día
                            MES=mes_seleccionado,
                            AÑO=año_seleccionado
                            ).first()
                        if registro_anterior:
                            producto.DIA_ANTERIOR = registro_anterior.DIA_ACTUAL
                            producto.DIA_ACTUAL = 0.0
                            producto.CAJAS = 0
                            producto.COMPRAS = 0.0
                            producto.ENTRADA = 0.0
                            producto.SALIDA = 0.0
                            producto.MERMAS = 0.0

        return render(request, 'cuentas/auditoria_4C.html', {
            'productos': productos,
            'dias': dias,
            'años': años,
            'dia_semanal_seleccionado': dia_semanal_seleccionado,
            'dia_seleccionado': dia_seleccionado,
            'mes_seleccionado': mes_seleccionado,
            'año_seleccionado': año_seleccionado,
        })

    # Redirigir de nuevo a la página de auditoria después de procesar los datos
    return redirect('cuentas:auditoria_4C')

def resumen_4C(request):
     # Generar una lista de días del 1 al 31
    dias = list(range(1, 32))
    # Generar una lista de años desde 2020 hasta 2050
    años = list(range(2023, 2051))

    if request.method == 'POST':
        mes_seleccionado = request.POST.get('mes', '')
        año_seleccionado = int(request.POST.get('anio', 0))
        dia_inicial = int(request.POST.get('dia_inicial', 0))
        dia_final = int(request.POST.get('dia_final', 0))
        
        registros = Registro_Auditorias_4C.objects.filter(MES=mes_seleccionado, AÑO=año_seleccionado)
        registros2 = registros.filter(DIA__gte=dia_inicial, DIA__lte=dia_final)

        # Crear un diccionario para almacenar los registros por DESCRIPCION
        registros_por_descripcion = defaultdict(list)
        for registro in registros:
            descripcion = registro.DESCRIPCION
            registros_por_descripcion[descripcion].append(registro)

        registros_agrupados = []
        for descripcion, registros in registros_por_descripcion.items():
            registro_agrupado = {
                'descripcion': descripcion,
                'medida': registros[0].MEDIDA,  # Tomamos la MEDIDA del primer registro
                'consumos_diarios': [registro.CONSUMO_DIARIO for registro in registros]
            }
            registros_agrupados.append(registro_agrupado)
        
        registros_por_dia = {}
        for registro in registros:
            dia = registro.DIA
            if dia not in registros_por_dia:
                registros_por_dia[dia] = []
                registros_por_dia[dia].append(registro)

        registros_agrupados2 = []  # Inicializa la lista fuera del bucle
        for descripcion, registros in registros_por_descripcion.items():
            consumos_diarios = [registro.CONSUMO_DIARIO for registro in registros if registro in registros2]
            promedio = sum(consumos_diarios) / len(consumos_diarios) if consumos_diarios else 0
            suma = sum(consumos_diarios)
            promedio_redondeado = round(promedio, 3)
            suma_redondeada = round(suma,3)
            registro_agrupado = {
                'descripcion': descripcion,
                'medida': registros[0].MEDIDA,
                'consumos_diarios': consumos_diarios,
                'promedio': promedio_redondeado,
                'suma': suma_redondeada,
            }
            registros_agrupados2.append(registro_agrupado)
            

        return render(request, 'cuentas/resumen_4C.html', {
            'dias': dias,
            'años': años,
            'registros_agrupados': registros_agrupados,
            'mes_seleccionado': mes_seleccionado,
            'año_seleccionado': año_seleccionado,
            'registros_por_dia': registros_por_dia,
            'registros_agrupados2': registros_agrupados2,
            'dia_inicial': dia_inicial,  # Pasar los valores de día inicial y final
            'dia_final': dia_final,
        })

    return render(request, 'cuentas/resumen_4C.html', {'dias':dias,'años': años})

def auditoria_4C_Bebidas(request):
    productos = Auditorias_4C_Bebidas.objects.all()

    # Generar una lista de días del 1 al 31
    dias = list(range(1, 32))

    # Generar una lista de años desde 2020 hasta 2050
    años = list(range(2023, 2051))

    if request.method == 'POST' and 'limpiar_Bebidas_4C' in request.POST:

        for producto in productos:
            producto.DIA = 0
            producto.MES = 'Mes'
            producto.AÑO = 0 
            producto.DIA_SEMANAL = 'Dia'
            producto.DIA_ANTERIOR = 0
            producto.DIA_ACTUAL = 0
            producto.COMPRAS = 0
            producto.ENTRADA = 0
            producto.SALIDA = 0
            producto.CONSUMO_DIARIO = 0
            producto.save()
        return render(request, 'cuentas/auditoria_4C_Bebidas.html', {'productos': productos, 'dias': dias, 'años': años})
    
    if request.method == 'POST':
        if 'archivo' in request.FILES:
            archivo = request.FILES['archivo']
            workbook = openpyxl.load_workbook(archivo)
            sheet = workbook.active

            productos_nuevos = []

            for row in sheet.iter_rows(min_row=4, values_only=True):
                DESCRIPCION, MEDIDA = row[1], str(row[2])

                if DESCRIPCION and MEDIDA:
                    producto, created = Auditorias_4C_Bebidas.objects.get_or_create(
                        DESCRIPCION=DESCRIPCION,
                        defaults={
                            'MEDIDA': MEDIDA,
                            'DIA_ANTERIOR': 0,
                            'DIA_ACTUAL': 0,
                            'COMPRAS': 0,
                            'ENTRADA': 0,
                            'SALIDA': 0,
                            'CONSUMO_DIARIO': 0,
                            'DIA_SEMANAL': 'Dia',
                            'DIA': 0,
                            'MES': 'Mes',
                            'AÑO': 0,
                        }
                    )
                    if created:
                        productos_nuevos.append(producto)
                    else:
                        producto.DESCRIPCION = DESCRIPCION
                        producto.MEDIDA = MEDIDA
                        producto.save()

    if request.method == 'POST' and 'guardar_cambios' in request.POST:
        # Procesar la selección de DÍA, MES y AÑO
        dia_seleccionado = int(request.POST.get('dia', 0))
        mes_seleccionado = request.POST.get('mes', '')
        año_seleccionado = int(request.POST.get('anio', 0))
        dia_semanal_seleccionado = request.POST.get('dia_semanal', '')

        # Iterar sobre los productos y actualizar los campos correspondientes
        for producto in productos:
            dia_anterior = int(request.POST.get(f'dia_anterior_{producto.id}', 0))
            dia_actual = int(request.POST.get(f'dia_actual_{producto.id}', 0))
            compras = int(request.POST.get(f'compras_{producto.id}', 0))
            entrada = int(request.POST.get(f'entrada_{producto.id}', 0))
            salida = int(request.POST.get(f'salida_{producto.id}', 0))

            # Actualizar los campos del producto
            producto.DIA = dia_seleccionado
            producto.MES = mes_seleccionado
            producto.AÑO = año_seleccionado
            producto.DIA_SEMANAL = dia_semanal_seleccionado
            producto.DIA_ANTERIOR = dia_anterior
            producto.DIA_ACTUAL = dia_actual
            producto.COMPRAS = compras
            producto.ENTRADA = entrada
            producto.SALIDA = salida

            # Calcular el consumo diario
            if dia_actual == 0:
                producto.CONSUMO_DIARIO = 0
            else:
                producto.CONSUMO_DIARIO = (-dia_anterior) - compras + dia_actual - entrada + salida 

            producto.save()

                # Verificar si existe un registro en Registro_Auditorias_HEB con la misma fecha
            registro_existente = Registro_Auditorias_4C_Bebidas.objects.filter(
                DESCRIPCION=producto.DESCRIPCION,
                MEDIDA=producto.MEDIDA,
                DIA=dia_seleccionado,
                MES=mes_seleccionado,
                AÑO=año_seleccionado
            ).first()

            if registro_existente:
                 # Si existe un registro con la misma fecha, actualizar los datos en Registro_Auditorias_HEB
                registro_existente.DIA_ANTERIOR = producto.DIA_ANTERIOR
                registro_existente.DIA_ACTUAL = producto.DIA_ACTUAL
                registro_existente.COMPRAS = producto.COMPRAS
                registro_existente.ENTRADA = producto.ENTRADA
                registro_existente.SALIDA = producto.SALIDA
                registro_existente.CONSUMO_DIARIO = producto.CONSUMO_DIARIO
                registro_existente.save()
            else:
            # Si no existe un registro con la misma fecha, insertar un nuevo registro en Registro_Auditorias_HEB
                nuevo_registro = Registro_Auditorias_4C_Bebidas(
                    DESCRIPCION=producto.DESCRIPCION,
                    MEDIDA=producto.MEDIDA,
                    DIA_ANTERIOR=producto.DIA_ANTERIOR,
                    DIA_ACTUAL=producto.DIA_ACTUAL,
                    COMPRAS=producto.COMPRAS,
                    ENTRADA=producto.ENTRADA,
                    SALIDA=producto.SALIDA,
                    CONSUMO_DIARIO=producto.CONSUMO_DIARIO,
                    DIA_SEMANAL = dia_semanal_seleccionado,
                    DIA=dia_seleccionado,
                    MES=mes_seleccionado,
                    AÑO=año_seleccionado
                    )
                nuevo_registro.save()
                
        return render(request, 'cuentas/auditoria_4C_Bebidas.html', {
            'dias': dias,
            'años': años,
            'productos': productos,
            'dia_semanal_seleccionado': dia_semanal_seleccionado,
            'dia_seleccionado': dia_seleccionado,
            'mes_seleccionado': mes_seleccionado,
            'año_seleccionado': año_seleccionado,
        })

    return render(request, 'cuentas/auditoria_4C_Bebidas.html', {'productos': productos, 'dias': dias, 'años': años})

def recuperar_datos_4C_Bebidas(request):
    # Generar una lista de días del 1 al 31
    dias = list(range(1, 32))

    # Generar una lista de años desde 2020 hasta 2050
    años = list(range(2023, 2051))

    if request.method == 'POST':
        # Obtener el día, mes y año seleccionados
        dia_seleccionado = int(request.POST.get('dia', 0))
        mes_seleccionado = request.POST.get('mes', '')
        año_seleccionado = int(request.POST.get('anio', 0))
        dia_semanal_seleccionado = request.POST.get('dia_semanal', '')
        
        productos = Auditorias_4C_Bebidas.objects.all()
        
        for producto in productos:
            registro = Registro_Auditorias_4C_Bebidas.objects.filter(
                DESCRIPCION=producto.DESCRIPCION,
                MEDIDA=producto.MEDIDA,
                DIA=dia_seleccionado,
                MES=mes_seleccionado,
                AÑO=año_seleccionado
            ).first()
            
            if registro:
                # Si existen registros, actualizar los datos en Auditorias_HEB
                producto.DIA_ANTERIOR = registro.DIA_ANTERIOR
                producto.DIA_ACTUAL = registro.DIA_ACTUAL
                producto.COMPRAS = registro.COMPRAS
                producto.ENTRADA = registro.ENTRADA
                producto.SALIDA = registro.SALIDA
                producto.CONSUMO_DIARIO = registro.CONSUMO_DIARIO
                producto.DIA_SEMANAL = registro.DIA_SEMANAL
                producto.DIA = registro.DIA
                producto.MES = registro.MES
                producto.AÑO = registro.AÑO
                producto.save()
            else:
                if dia_seleccionado == 1 and año_seleccionado == 2023:
                    if mes_seleccionado == 'Septiembre':
                        registro_anterior = Registro_Auditorias_4C_Bebidas.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=31, 
                            MES='Agosto',
                            AÑO=año_seleccionado
                            ).first()
                    elif mes_seleccionado == 'Octubre':
                        registro_anterior = Registro_Auditorias_4C_Bebidas.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=30, 
                            MES='Septiembre',
                            AÑO=año_seleccionado
                            ).first()
                    elif mes_seleccionado == 'Noviembre':
                        registro_anterior = Registro_Auditorias_4C_Bebidas.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=31, 
                            MES='Octubre',
                            AÑO=año_seleccionado
                            ).first()
                    elif mes_seleccionado == 'Diciembre':
                        registro_anterior = Registro_Auditorias_4C_Bebidas.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=30, 
                            MES='Noviembre',
                            AÑO=año_seleccionado
                            ).first()
                    if registro_anterior:
                        producto.DIA_ANTERIOR = registro_anterior.DIA_ACTUAL
                        producto.DIA_ACTUAL = 0
                        producto.COMPRAS = 0
                        producto.ENTRADA = 0
                        producto.SALIDA = 0
                        producto.save()
                else:
                    if dia_seleccionado == 1 and año_seleccionado == 2024:
                        if mes_seleccionado == 'Enero':
                            registro_anterior = Registro_Auditorias_4C_Bebidas.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Diciembre',
                                AÑO=2023
                                ).first()
                        if mes_seleccionado == 'Febrero':
                            registro_anterior = Registro_Auditorias_4C_Bebidas.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Enero',
                                AÑO=año_seleccionado
                                ).first()
                        elif mes_seleccionado == 'Marzo':
                            registro_anterior = Registro_Auditorias_4C_Bebidas.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=28, 
                                MES='Febrero',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Abril':
                            registro_anterior = Registro_Auditorias_4C_Bebidas.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=31, 
                            MES='Marzo',
                            AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Mayo':
                            registro_anterior = Registro_Auditorias_4C_Bebidas.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=30, 
                                MES='Abril',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Junio':
                            registro_anterior = Registro_Auditorias_4C_Bebidas.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Mayo',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Julio':
                            registro_anterior = Registro_Auditorias_4C_Bebidas.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=30, 
                                MES='Junio',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Agosto':
                            registro_anterior = Registro_Auditorias_4C_Bebidas.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Julio',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Septiembre':
                            registro_anterior = Registro_Auditorias_4C_Bebidas.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Agosto',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Octubre':
                            registro_anterior = Registro_Auditorias_4C_Bebidas.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=30, 
                                MES='Septiembre',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Noviembre':
                            registro_anterior = Registro_Auditorias_4C_Bebidas.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=31, 
                                MES='Octubre',
                                AÑO=año_seleccionado
                            ).first()
                        elif mes_seleccionado == 'Diciembre':
                            registro_anterior = Registro_Auditorias_4C_Bebidas.objects.filter(
                                DESCRIPCION=producto.DESCRIPCION,
                                MEDIDA=producto.MEDIDA,
                                DIA=30, 
                                MES='Noviembre',
                                AÑO=año_seleccionado
                            ).first()
                        if registro_anterior:
                            producto.DIA_ANTERIOR = registro_anterior.DIA_ACTUAL
                            producto.DIA_ACTUAL = 0
                            producto.COMPRAS = 0
                            producto.ENTRADA = 0
                            producto.SALIDA = 0
                            producto.save()
                    # Si no existe un registro, busca el registro del día anterior (DIA - 1)
                    else: 
                        registro_anterior = Registro_Auditorias_4C_Bebidas.objects.filter(
                            DESCRIPCION=producto.DESCRIPCION,
                            MEDIDA=producto.MEDIDA,
                            DIA=dia_seleccionado - 1,  # Restar 1 al día
                            MES=mes_seleccionado,
                            AÑO=año_seleccionado
                            ).first()
                        if registro_anterior:
                            producto.DIA_ANTERIOR = registro_anterior.DIA_ACTUAL
                            producto.DIA_ACTUAL = 0
                            producto.COMPRAS = 0
                            producto.ENTRADA = 0
                            producto.SALIDA = 0
                            producto.save()

        return render(request, 'cuentas/auditoria_4C_Bebidas.html', {
            'productos': productos,
            'dias': dias,
            'años': años,
            'dia_semanal_seleccionado': dia_semanal_seleccionado,
            'dia_seleccionado': dia_seleccionado,
            'mes_seleccionado': mes_seleccionado,
            'año_seleccionado': año_seleccionado,
        })

    # Redirigir de nuevo a la página de auditoria después de procesar los datos
    return redirect('cuentas:auditoria_4C_Bebidas')

def resumen_4C_Bebidas(request):
    # Generar una lista de días del 1 al 31
    dias = list(range(1, 32))
    # Generar una lista de años desde 2020 hasta 2050
    años = list(range(2023, 2051))

    if request.method == 'POST':
        mes_seleccionado = request.POST.get('mes', '')
        año_seleccionado = int(request.POST.get('anio', 0))
        dia_inicial = int(request.POST.get('dia_inicial', 0))
        dia_final = int(request.POST.get('dia_final', 0))
        
        registros = Registro_Auditorias_4C_Bebidas.objects.filter(MES=mes_seleccionado, AÑO=año_seleccionado)
        registros2 = registros.filter(DIA__gte=dia_inicial, DIA__lte=dia_final)

        # Crear un diccionario para almacenar los registros por DESCRIPCION
        registros_por_descripcion = defaultdict(list)
        for registro in registros:
            descripcion = registro.DESCRIPCION
            registros_por_descripcion[descripcion].append(registro)

        registros_agrupados = []
        for descripcion, registros in registros_por_descripcion.items():
            registro_agrupado = {
                'descripcion': descripcion,
                'medida': registros[0].MEDIDA,  # Tomamos la MEDIDA del primer registro
                'consumos_diarios': [registro.CONSUMO_DIARIO for registro in registros]
            }
            registros_agrupados.append(registro_agrupado)
        
        registros_por_dia = {}
        for registro in registros:
            dia = registro.DIA
            if dia not in registros_por_dia:
                registros_por_dia[dia] = []
                registros_por_dia[dia].append(registro)

        registros_agrupados2 = []  # Inicializa la lista fuera del bucle
        for descripcion, registros in registros_por_descripcion.items():
            consumos_diarios = [registro.CONSUMO_DIARIO for registro in registros if registro in registros2]
            promedio = sum(consumos_diarios) / len(consumos_diarios) if consumos_diarios else 0
            suma = sum(consumos_diarios)
            promedio_redondeado = round(promedio, 3)
            suma_redondeada = round(suma,3)
            registro_agrupado = {
                'descripcion': descripcion,
                'medida': registros[0].MEDIDA,
                'consumos_diarios': consumos_diarios,
                'promedio': promedio_redondeado,
                'suma': suma_redondeada,
            }
            registros_agrupados2.append(registro_agrupado)
            

        return render(request, 'cuentas/resumen_4C_Bebidas.html', {
            'dias': dias,
            'años': años,
            'registros_agrupados': registros_agrupados,
            'mes_seleccionado': mes_seleccionado,
            'año_seleccionado': año_seleccionado,
            'registros_por_dia': registros_por_dia,
            'registros_agrupados2': registros_agrupados2,
            'dia_inicial': dia_inicial,  # Pasar los valores de día inicial y final
            'dia_final': dia_final,
        })

    return render(request, 'cuentas/resumen_4C_Bebidas.html', {'dias':dias,'años': años})
